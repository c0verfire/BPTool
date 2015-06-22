#!/usr/bin/python

import elementtree.ElementTree as ET
from sys import argv
import os
import csv
import requests
import re
import time
from openpyxl import Workbook

script, fwinfo = argv

wb = Workbook()
ws = wb.active
ws['A1'] = 'IP Address/FQDN'
ws['B1'] = 'BP Number'
ws['C1'] = 'Title'
ws['D1'] = 'Priority'
ws['E1'] = 'Status'
ws['F1'] = 'Description'

datetime = time.strftime("%Y-%m-%d %H%M")


#-------Rule Definitions------
#----Rule 01000 - 03999: Global Rules
def BP01000(ip, apikey):
	bpnum = "BP01000"
	title = "Configure Hostname on System"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	rxpath = "result/hostname"
	xpath = "/config/devices/entry[@name='localhost.localdomain']/deviceconfig/system/hostname"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	rresp = ET.fromstring(rrule.content)
	result = rresp.find(rxpath)
	if re.match("Panorama", result.text):
		status = "Fail"
		mesg = "%s Default Panorama Hostname" % ip
	elif re.match("PA-\d{3,4}", result.text):
		status = "Fail"
		mesg = "%s Default PAN Firewall Hostname" % ip
	elif re.match("PA-VM", result.text):
		status = "Fail"
		mesg = "%s Default PAN Firewall Hostname" % ip
	else:
		status = "Pass"
		mesg = "Hostnames are Defined"
		
	ws.append([ip, bpnum, title, priority, status, mesg])
	
def BP01001(ip, apikey):
	bpnum = "BP01001"
	title = "Configure Domain Name on System"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	rxpath = "result/domain"
	xpath = "/config/devices/entry[@name='localhost.localdomain']/deviceconfig/system/domain"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	rresp = ET.fromstring(rrule.content)
	result = rresp.find(rxpath)
	if (result is None):
		status = "Fail"
		mesg = "%s Domain Name not Defined" % ip
	elif re.match(".+", result.text):
		status = "Pass"
		mesg = "%s Domain Name Defined" % ip
	else:
		print "FAIL!!!"
		
	ws.append([ip, bpnum, title, priority, status, mesg])
	
def BP01002(ip, apikey):
	bpnum = "BP01002"
	title = "Replace default administrative super-user account by creating a new superuser account and then deleting the admin account."
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	xpath = "/config/mgt-config/users"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	
	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall("./result/users/entry"):
		username = entryElement.attrib['name']
		for rolebasedElement in entryElement.findall('permissions/role-based'):
			superuserElement = rolebasedElement.find('superuser')
			if not superuserElement is None and superuserElement.text.lower() == 'yes' and username == 'admin':
				status = "Fail"
				mesg = "User Admin is Present"
				ws.append([ip, bpnum, title, priority, status, mesg])
			if not superuserElement is None and superuserElement.text.lower() == 'yes' and username != 'admin':
				status = "Pass"
				mesg = username + ' is superuser'
				ws.append([ip, bpnum, title, priority, status, mesg])
			superreaderElement = rolebasedElement.find('superreader')
			if not superreaderElement is None and superreaderElement.text.lower() == 'yes':
				status = "Informational"
				mesg = username + ' is superreader'
				ws.append([ip, bpnum, title, priority, status, mesg])

def BP01003(ip, apikey):
	bpnum = "BP01003"
	title = "Configure administrative access lockouts and timeouts in the authentication settings portion of the device setup tab. Configure administrative timeout to 10 minutes; apply these settings to administrator accounts."
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	rxpath = "result/idle-timeout"
	xpath = "/config/devices/entry[@name='localhost.localdomain']/deviceconfig/setting/management/idle-timeout"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	rresp = ET.fromstring(rrule.content)
	result = rresp.find(rxpath)
	if (result is None):
		status = "Fail"
		mesg = "Idle Timeout should be set to 10 minutes"
	elif result.text != "10":
		status = "Fail"
		mesg = "Idle Timeout should be set to 10 minutes"
	elif result.text == "10":
		status = "Pass"
		mesg = "Idle Timeout is set to 10 minutes"
		
	ws.append([ip, bpnum, title, priority, status, mesg])

def BP01004(ip, apikey):
	bpnum = "BP01004"
	title = "Configure Local User Authentication Profile"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Authentication Profile
	xpath = "/config/shared/authentication-profile"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	# Query for User Information
	xpath2 = "/config/mgt-config/users"
	rulequery2 = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath2}
	rrule2 = requests.get('https://' + ip + '/api', params = rulequery2, verify=False)
	
	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall("./result/authentication-profile/entry"):
		authprof = entryElement.attrib['name']
		for methodElement in entryElement.findall('method'):
			localdbElement = methodElement.find('local-database')
			for lockoutElement in entryElement.findall('lockout'):
				timeElement = lockoutElement.find('lockout-time')
				failElement = lockoutElement.find('failed-attempts')
				goodauth = []
				if not localdbElement is None and timeElement.text.lower() == '15' and failElement.text.lower() == '3':
					goodauth.append(authprof)
					for a in goodauth:
						status = "Pass"
						mesg = a + ' is configured correctly'
						ws.append([ip, bpnum, title, priority, status, mesg])

						responseElement2 = ET.fromstring(rrule2.text)
						usrgrp = {}
						for userElement in responseElement2.findall("./result/users/entry"):
							username = userElement.attrib['name']
							authprofElement = userElement.find('authentication-profile')
							if authprofElement is None:
								status = "Informational"
								mesg = username + ' does not have an Authentication Group assigned'
								ws.append([ip, bpnum, title, priority, status, mesg])										
							if not authprofElement is None:
								usrgrp[username] = authprofElement.text
								for username, authprofElement in usrgrp.iteritems():
									if authprofElement in goodauth:
										status = "Pass"
										mesg = username + ' is using Authentication Group ' + authprofElement
										ws.append([ip, bpnum, title, priority, status, mesg])
									

def BP01005(ip, apikey):
	bpnum = "BP01005"
	title = "Configure the firewall to verify the identity of the Palo Alto update server on the device setup page."
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	rxpath = "result/system/server-verification"
	xpath = "/config/devices/entry[@name='localhost.localdomain']/deviceconfig/system"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	rresp = ET.fromstring(rrule.content)
	result = rresp.find(rxpath)
	if result is None:
		status = "Fail"
		mesg = "Server Verification Should be Enabled"
		ws.append([ip, bpnum, title, priority, status, mesg])
	elif result.text == "yes":
		status = "Pass"
		mesg = "Server Verification is Enabled"	
		ws.append([ip, bpnum, title, priority, status, mesg])	
	elif result.text == "no":
		status = "Fail"
		mesg = "Server Verification Should be Enabled"
		ws.append([ip, bpnum, title, priority, status, mesg])
	
	
def BP01006(ip, apikey):
	bpnum = "BP01006"
	title = "Configure login banner."
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	rxpath = "result/system/login-banner"
	xpath = "/config/devices/entry[@name='localhost.localdomain']/deviceconfig/system"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	rresp = ET.fromstring(rrule.content)
	result = rresp.find(rxpath)
	if (result is None):
		status = "Fail"
		mesg = "Login Banner Should Be Configured"
	elif (result is not None):
		status = "Pass"
		mesg = "Login Banner Is Configured"	
		
	ws.append([ip, bpnum, title, priority, status, mesg])	


def BP01007(ip, apikey):
	bpnum = "BP01007"
	title = "Configure Geo-Location Longitude and Latitude"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	rxpath = "result/system/geo-location"
	xpath = "/config/devices/entry[@name='localhost.localdomain']/deviceconfig/system"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	rresp = ET.fromstring(rrule.content)
	result = rresp.find(rxpath)
	if (result is None):
		status = "Fail"
		mesg = "Geolocation Coordinates should be configured"
	elif (result is not None):
		status = "Pass"
		mesg = "Geolocation Coordinates are configured"	
		
	ws.append([ip, bpnum, title, priority, status, mesg])	
	

def BP01008(ip, apikey):
	bpnum = "BP01008"
	title = "Configure the firewall to use redundant DNS Servers"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	xpath = "/config/devices/entry[@name='localhost.localdomain']/deviceconfig/system"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)

	responseElement = ET.fromstring(rrule.text)
	for srvElement in responseElement.findall("./result/system/dns-setting/servers"):
		PrimarySrvElement = srvElement.find('primary')
		SecondarySrvElement = srvElement.find('secondary')
		if PrimarySrvElement is None: 
			status = "Fail"
			mesg = "DNS Servers are not configured"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif not PrimarySrvElement is None and SecondarySrvElement is None:
			status = "Fail"
			mesg = "Only Primary DNS Server Configured. Primary Server IP: %s" % PrimarySrvElement.text
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif not PrimarySrvElement is None and not SecondarySrvElement is None:
			status = "Pass"
			mesg = "Primary and Secondary DNS Servers Configured. Primary Server IP: %s Secondary Server IP: %s" % (PrimarySrvElement.text, SecondarySrvElement.text)
			ws.append([ip, bpnum, title, priority, status, mesg])
			
def BP01009(ip, apikey):
	bpnum = "BP01009"
	title = "Configure the firewall to use redundant NTP Servers"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	xpath = "/config/devices/entry[@name='localhost.localdomain']/deviceconfig/system"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	
	responseElement = ET.fromstring(rrule.text)
	for psrvElement in responseElement.findall("./result/system/ntp-servers/primary-ntp-server"):
		PrimarySrvElement = psrvElement.find('ntp-server-address')
		for noneElement in psrvElement.findall('authentication-type/none'):
			status = "Informational"
			mesg = "NTP Server %s configured with no authentication" % PrimarySrvElement.text
			ws.append([ip, bpnum, title, priority, status, mesg])
		for authElement in psrvElement.findall('authentication-type/symmetric-key/algorithm/*'):
			if authElement.tag == 'md5':
				status = "Fail"
				mesg = "Authencation type md5 found on NTP Server %s" % PrimarySrvElement.text
				ws.append([ip, bpnum, title, priority, status, mesg])
			if authElement.tag == 'sha1':
				status = "Pass"
				mesg = "Authencation type sha1 found on NTP Server %s" % PrimarySrvElement.text
				ws.append([ip, bpnum, title, priority, status, mesg])
				
	for ssrvElement in responseElement.findall("./result/system/ntp-servers/secondary-ntp-server"):
		SecondarySrvElement = ssrvElement.find('ntp-server-address')
		for noneElement in ssrvElement.findall('authentication-type/none'):
			status = "Informational"
			mesg = "NTP Server %s configured with no authentication" % SecondarySrvElement.text
			ws.append([ip, bpnum, title, priority, status, mesg])
		for authElement in ssrvElement.findall('authentication-type/symmetric-key/algorithm/*'):
			if authElement.tag == 'md5':
				status = "Fail"
				mesg = "Authencation type md5 found on NTP Server %s" % SecondarySrvElement.text
				ws.append([ip, bpnum, title, priority, status, mesg])
			if authElement.tag == 'sha1':
				status = "Pass"
				mesg = "Authencation type sha1 found on NTP Server %s" % SecondarySrvElement.text
				ws.append([ip, bpnum, title, priority, status, mesg])	
				
def BP01010(ip, apikey):
	bpnum = "BP01010"
	title = "Limit management interface traffic to ping and secure protocols only"
	priority = "High"
	print "Running Rule %s - %s" % (bpnum, title)
	xpath = "/config/devices/entry[@name='localhost.localdomain']/deviceconfig/system"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	
	responseElement = ET.fromstring(rrule.text)
	for srvElement in responseElement.findall("./result/system/service"):
		ssh = srvElement.find('disable-ssh')
		telnet = srvElement.find('disable-telnet')
		http = srvElement.find('disable-http')
		https = srvElement.find('disable-https')
		httpocsp = srvElement.find('disable-http-ocsp')
		snmp = srvElement.find('disable-snmp')
		icmp = srvElement.find('disable-icmp')
		uid = srvElement.find('disable-userid-service')
		uidlogssl = srvElement.find('disable-userid-syslog-listener-ssl')
		uidlogudp = srvElement.find('disable-userid-syslog-listener-udp')

		if ssh is None:
			status = "Pass"
			mesg = "SSH is Enabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif ssh.text == 'yes':
			status = "Fail"
			mesg = "SSH is Disabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif ssh.text == 'no':
			status = "Pass"
			mesg = "SSH is Enabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
			
		if telnet.text == 'yes':
			status = "Pass"
			mesg = "Telnet is Disabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif telnet.text == 'no':
			status = "Fail"
			status = "Telnet is Enabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		
		if http.text == 'yes':
			status = "Pass"
			mesg = "HTTP is Disabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif http.text == 'no':
			status = "Fail"
			status = "HTTP is Enabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
			
		if https is None:
			status = "Pass"
			mesg = "HTTPS is Enabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif https.text == 'yes':
			status = "Fail"
			mesg = "HTTPS is Disabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif https.text == 'no':
			status = "Pass"
			mesg = "HTTPS is Enabled"
			ws.append([ip, bpnum, title, priority, status, mesg])	
		
		if httpocsp is None:
			status = "Informational"
			mesg = "HTTP OSCP is Disabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif httpocsp.text == 'yes':
			status = "Informational"
			mesg = "HTTP OSCP is Disabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif httpocsp.text == 'no':
			status = "Informational"
			mesg = "HTTP OSCP is Enabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
			
		if snmp is None:
			status = "Informational"
			mesg = "SNMP Polling is Disabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif snmp.text == 'yes':
			status = "Informational"
			mesg = "SNMP Polling is Disabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif snmp.text == 'no':
			status = "Informational"
			mesg = "SNMP Polling is Enabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
			
		if icmp is None:
			status = "Informational"
			mesg = "ICMP is Disabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif icmp.text == 'yes':
			status = "Informational"
			mesg = "ICMP is Disabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif icmp.text == 'no':
			status = "Informational"
			mesg = "ICMP is Enabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
			
		if uid is None:
			status = "Informational"
			mesg = "User-ID Polling is Disabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif uid.text == 'yes':
			status = "Informational"
			mesg = "User-ID Polling is Disabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif uid.text == 'no':
			status = "Informational"
			mesg = "User-ID Polling is Enabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		
		if uidlogssl is None:
			status = "Informational"
			mesg = "User SSL Log Polling is Disabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif uidlogssl.text == 'yes':
			status = "Informational"
			mesg = "User SSL Log Polling is Disabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif uidlogssl.text == 'no':
			status = "Informational"
			mesg = "User SSL Log Polling is Enabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
			
		if uidlogudp is None:
			status = "Informational"
			mesg = "User UDP Log Polling is Disabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif uidlogudp.text == 'yes':
			status = "Informational"
			mesg = "User UDP Log Polling is Disabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif uidlogudp.text == 'no':
			status = "Informational"
			mesg = "User UDP Log Polling is Enabled"
			ws.append([ip, bpnum, title, priority, status, mesg])

def BP01011(ip, apikey):
	bpnum = "BP01011"
	title = "Limit permitted IP Addresses to those necessary for device management"
	priority = "High"
	print "Running Rule %s - %s" % (bpnum, title)
	xpath = "/config/devices/entry[@name='localhost.localdomain']/deviceconfig/system"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	
	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall("./result/system"):
		entry = entryElement.find('permitted-ip')
		if entry is None:
			status = "Fail"
			mesg = "No Management ACL has been defined."
			ws.append([ip, bpnum, title, priority, status, mesg])
	for permElement in responseElement.findall("./result/system/permitted-ip/entry"):
		mgmtip = permElement.attrib['name']
		if mgmtip == '0.0.0.0/0':
			status = "Fail"
			mesg = "Management IP ACL has Default Entry 0.0.0.0/0"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif mgmtip:
			status = "Pass"
			mesg = "Valid Management ACL has been defined for IP %s." % mgmtip
			ws.append([ip, bpnum, title, priority, status, mesg])
			
def BP01012(ip, apikey):
	bpnum = "BP01012"
	title = "Enable Log on High DP Load"
	priority = "High"
	print "Running Rule %s - %s" % (bpnum, title)
	xpath = "/config/devices/entry[@name='localhost.localdomain']/deviceconfig/setting"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	
	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall("./result/setting/management"):
		loghighdp = entryElement.find('enable-log-high-dp-load')
		if loghighdp is None:
			status = "Fail"
			mesg = "High DP Logging Not Enabled."
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif loghighdp.text == 'yes':
			status = "Pass"
			mesg = "High DP Logging Is Enabled."
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif loghighdp.text == 'no':
			status = "Fail"
			mesg = "High DP Logging Not Enabled."
			ws.append([ip, bpnum, title, priority, status, mesg])
			
def BP01013(ip, apikey):
	bpnum = "BP01013"
	title = "Secure Production Interface Management Profiles"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Interface Management Profile
	xpath = "/config/devices/entry[@name='localhost.localdomain']/network/profiles/interface-management-profile"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	# Query for Interface Information
	xpath2 = "/config/devices/entry[@name='localhost.localdomain']/network/interface/ethernet"
	rulequery2 = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath2}
	rrule2 = requests.get('https://' + ip + '/api', params = rulequery2, verify=False)
	
	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall("./result/interface-management-profile/entry"):
		profile = entryElement.attrib['name']
		ssh = entryElement.find('ssh')
		https = entryElement.find('https')
		ping = entryElement.find('ping')
		http = entryElement.find('http')
		telnet = entryElement.find('telnet')
		goodauth = []
		badauth = []
		if http is None and telnet is None:
			goodauth.append(profile)
			for a in goodauth:
				status = "Pass"
				mesg = a + ' is configured correctly'
				ws.append([ip, bpnum, title, priority, status, mesg])
		if not http is None and http.text == 'yes' or not telnet is None and telnet.text == 'yes':
			badauth.append(profile)
			for a in badauth:
				status = "Fail"
				mesg = a + ' is configured incorrectly'
				ws.append([ip, bpnum, title, priority, status, mesg])
		if not http is None and http.text == 'no' or not telnet is None and telnet.text == 'no':
			goodauth.append(profile)
			for a in goodauth:
				status = "Pass"
				mesg = a + ' is configured correctly'
				ws.append([ip, bpnum, title, priority, status, mesg])

		responseElement2 = ET.fromstring(rrule2.text)
		for intfElement in responseElement2.findall("./result/ethernet/entry"):
			intfgrp = {}
			interface = intfElement.attrib['name']
			mgmtprofElement = intfElement.find('layer3/interface-management-profile')
			if mgmtprofElement is None:
				status = "Informational"
				mesg = interface + ' does not have an Management Profile assigned'
				ws.append([ip, bpnum, title, priority, status, mesg])										
			if not mgmtprofElement is None and not interface in intfgrp:
				intfgrp[interface] = mgmtprofElement.text
				for interface, mgmtprofElement in intfgrp.iteritems():
					if mgmtprofElement in goodauth:
						status = "Pass"
						mesg = interface + ' is using Secure Management Profile ' + mgmtprofElement
						ws.append([ip, bpnum, title, priority, status, mesg])
					if mgmtprofElement in badauth:
						status = "Fail"
						mesg = interface + ' is using unsecure Management Profile ' + mgmtprofElement
						ws.append([ip, bpnum, title, priority, status, mesg])

def BP01014(ip, apikey):
	bpnum = "BP01014"
	title = "Restrict Production Interface Management Profiles Source Addresses"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Management Profile
	xpath = "/config/devices/entry[@name='localhost.localdomain']/network/profiles/interface-management-profile"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	
	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall("./result/interface-management-profile/entry"):
		profile = entryElement.attrib['name']
		permip = entryElement.find('permitted-ip')
		if permip is None:
			status = "Fail"
			mesg = "No Management ACL has been defined for " + profile
			ws.append([ip, bpnum, title, priority, status, mesg])
		for ipElement in entryElement.findall("permitted-ip/entry"):
			mgmtip = ipElement.attrib['name']
			if mgmtip == '0.0.0.0/0':
				status = "Fail"
				mesg = "Management IP ACL has Default Entry 0.0.0.0/0 in Profile %s" % profile
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif mgmtip:
				status = "Pass"
				mesg = "Valid Management ACL has been defined for IP %s. in Profile %s" % (mgmtip, profile)
				ws.append([ip, bpnum, title, priority, status, mesg])

def BP01015(ip, apikey):
	bpnum = "BP01015"
	title = "Configure minimum password complexity profile "
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Management Profile
	xpath = "/config/mgt-config/password-complexity"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	
	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall("./result"):
		passcomp = entryElement.find('password-complexity')
		if passcomp is None:
			status = "Fail"
			mesg = "Password Complexity is not Enabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		for passcompElement in entryElement.findall('password-complexity'):
			enabled = passcompElement.find('enabled')
			minlen = passcompElement.find('minimum-length')
			minup = passcompElement.find('minimum-uppercase-letters')
			minlow = passcompElement.find('minimum-lowercase-letters')
			minnum = passcompElement.find('minimum-numeric-letters')
			minspe = passcompElement.find('minimum-special-characters')
			blkusr = passcompElement.find('block-username-inclusion')
			chrdif = passcompElement.find('new-password-differs-by-characters')
			passchg = passcompElement.find('password-change-on-first-login')
			passhis = passcompElement.find('password-history-count')
			expper = passcompElement.find('password-change/expiration-period')
			expwrn = passcompElement.find('password-change/expiration-warning-period')
			expgrc = passcompElement.find('password-change/post-expiration-grace-period')
			
			# Check if Password Complexity is Enabled
			if enabled.text == 'yes':
				status = "Pass"
				mesg = "Password Complexity is Enabled"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif enabled.text == 'no':
				status = "Fail"
				mesg = "Password Complexity is not Enabled"
				ws.append([ip, bpnum, title, priority, status, mesg])

			#Check if Password Minimum Length is 12 or greater
			if minlen is None:
				status = "Fail"
				mesg = "Password Minimum Length is not Configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif minlen.text >= '12':
				status = "Pass"
				mesg = "Password Minimum Length is 12 or greater"
				ws.append([ip, bpnum, title, priority, status, mesg])
			else:
				status = "Fail"
				mesg = "Password Minimum Length is not 12 or greater"
				ws.append([ip, bpnum, title, priority, status, mesg])
			
			#Check if Password Minimum Uppercase Characters is 1 or greater
			if minup is None:
				status = "Fail"
				mesg = "Password Minimum Uppercase Characters is not Configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif minup.text >= '1':
				status = "Pass"
				mesg = "Password Minimum Uppercase Characters is 1 or greater"
				ws.append([ip, bpnum, title, priority, status, mesg])
			else:
				status = "Fail"
				mesg = "Password Minimum Uppercase Characters is not 1 or greater"
				ws.append([ip, bpnum, title, priority, status, mesg])
			
			#Check if Password Minimum Lowercase Characters is 1 or greater
			if minlow is None:
				status = "Fail"
				mesg = "Password Minimum Lowercase Characters is not Configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif minlow.text >= '1':
				status = "Pass"
				mesg = "Password Minimum Lowercase Characters is 1 or greater"
				ws.append([ip, bpnum, title, priority, status, mesg])
			else:
				status = "Fail"
				mesg = "Password Minimum Lowercase Characters is not 1 or greater"
				ws.append([ip, bpnum, title, priority, status, mesg])

			#Check if Password Minimum Numeric Characters is 1 or greater
			if minnum is None:
				status = "Fail"
				mesg = "Password Minimum Numeric Characters is not Configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif minnum.text >= '1':
				status = "Pass"
				mesg = "Password Minimum Numeric Characters is 1 or greater"
				ws.append([ip, bpnum, title, priority, status, mesg])
			else:
				status = "Fail"
				mesg = "Password Minimum Numeric Characters is not 1 or greater"
				ws.append([ip, bpnum, title, priority, status, mesg])
				
			#Check if Password Minimum Special Characters is 1 or greater
			if minspe is None:
				status = "Fail"
				mesg = "Password Minimum Special Characters is not Configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif minspe.text >= '1':
				status = "Pass"
				mesg = "Password Minimum Special Characters is 1 or greater"
				ws.append([ip, bpnum, title, priority, status, mesg])
			else:
				status = "Fail"
				mesg = "Password Minimum Special Characters is not 1 or greater"
				ws.append([ip, bpnum, title, priority, status, mesg])
				
			#Check if Block Username Inclusion is  Enabled
			if blkusr is None:
				status = "Fail"
				mesg = "Block Username Inclusion is not Configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif blkusr.text == 'yes':
				status = "Pass"
				mesg = "Block Username Inclusion is Configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			else:
				status = "Fail"
				mesg = "Block Username Inclusion is not Configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
				
			#Check if New Password Differs by Three or More Characters
			if chrdif is None:
				status = "Fail"
				mesg = "Password Differentiation is not Configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif chrdif.text >= '3':
				status = "Pass"
				mesg = "Password Differentiation is configured by 3 characters or greater"
				ws.append([ip, bpnum, title, priority, status, mesg])
			else:
				status = "Fail"
				mesg = "Password Differentiation is not Configured"
				ws.append([ip, bpnum, title, priority, status, mesg])

			#Check if Force Password Change at First Login is Enabled
			if passchg is None:
				status = "Fail"
				mesg = "Force Password Change at First Login is not Configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif passchg.text == 'yes':
				status = "Pass"
				mesg = "Force Password Change at First Login is Configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			else:
				status = "Fail"
				mesg = "Force Password Change at First Login is not Configured"
				ws.append([ip, bpnum, title, priority, status, mesg])

			#Check if Password History of 24 or Greater is Enforced
			if passhis is None:
				status = "Fail"
				mesg = "Password History of 24 or greater is not configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif passhis.text >= '24':
				status = "Pass"
				mesg = "Password History of 24 or greater is not configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			else:
				status = "Fail"
				mesg = "Password History of 24 or greater is not configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
				
			#Check if Maximum of 90 Day Password Expiration is configured
			if expper is None:
				status = "Fail"
				mesg = "Password Expiration is not configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif expper.text <= '90':
				status = "Pass"
				mesg = "Maximum of 90 Day Password Expiration is configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			else:
				status = "Fail"
				mesg = "Maximum of 90 Day Password Expiration is not configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			
			#Check if Maximum of 7 Day Expiration Warning Period is configured
			if expwrn is None:
				status = "Fail"
				mesg = "Password Expiration Warning Period is not configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif expwrn.text <= '7':
				status = "Pass"
				mesg = "Maximum of 7 Day Password Expiration Warning Period is configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			else:
				status = "Fail"
				mesg = "Maximum of 7 Day Password Expiration Warning Period is not configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
				
			#Check if Maximum of 7 Day Post Expiration Grace Period is configured
			if expgrc is None:
				status = "Fail"
				mesg = "Post Expiration Grace Period is not configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif expgrc.text <= '7':
				status = "Pass"
				mesg = "Maximum of 7 Day Post Expiration Grace Period is configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			else:
				status = "Fail"
				mesg = "Maximum of 7 Day Post Expiration Grace Period is not configured"
				ws.append([ip, bpnum, title, priority, status, mesg])

def BP01016(ip, apikey):
	bpnum = "BP01016"
	title = "Firewall eTAC Recommended Versions of Code"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Management Profile
	xpath = "<show><system><info></info></system></show>"
	rulequery = {'type': 'op', 'action': 'get', 'key': apikey, 'cmd': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	
	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall("./result/system"):
		version = entryElement.find('sw-version')
		if version.text == "6.0.8" or version.text == "6.0.9" or version.text == "6.0.10" or version.text == "6.1.2" or version.text == "6.1.3" or version.text == "6.1.4":
			status = "Pass"
			mesg = "Current Version: %s - Firewall is running an eTAC Recommended Version of Code" % version.text
			ws.append([ip, bpnum, title, priority, status, mesg])
		else:
			status = "Fail"
			mesg = "Current Version: %s - Firewall is not running an eTAC Recommended Version of Code. For more information refer to https://intranet.paloaltonetworks.com/docs/DOC-4857" % version.text
			ws.append([ip, bpnum, title, priority, status, mesg])

def BP01017(ip, apikey):
	bpnum = "BP01017"
	title = "Panorama eTAC Recommended Versions of Code"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Management Profile
	xpath = "<show><system><info></info></system></show>"
	rulequery = {'type': 'op', 'action': 'get', 'key': apikey, 'cmd': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	
	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall("./result/system"):
		version = entryElement.find('sw-version')
		if version.text == "6.0.8" or version.text == "6.0.9" or version.text == "6.0.10" or version.text == "6.1.2" or version.text == "6.1.3" or version.text == "6.1.4":
			status = "Pass"
			mesg = "Current Version: %s - Firewall is running an eTAC Recommended Version of Code" % version.text
			ws.append([ip, bpnum, title, priority, status, mesg])
		else:
			status = "Fail"
			mesg = "Current Version: %s - Firewall is not running an eTAC Recommended Version of Code. For more information refer to https://intranet.paloaltonetworks.com/docs/DOC-4857" % version.text
			ws.append([ip, bpnum, title, priority, status, mesg])
			
#-------Rule Definitions------
#----Rule 04000 - 07999: Firewall Specific Rules
def BP04000(ip, apikey):
	bpnum = "BP04000"
	title = " Firewall Should Use Descriptive Rule Names"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Management Profile
	xpath = "/config/devices/entry[@name='localhost.localdomain']/vsys"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall("./result/vsys/entry"):
		vsys = entryElement.attrib['name']
		for secruleElement in entryElement.findall('rulebase/security/rules/entry'):
			rule = secruleElement.attrib['name']
			if re.match("rule\d{1,2}", rule):
				status = "Fail"
				mesg = "%s in %s is using a default rule name" % (rule, vsys)
				ws.append([ip, bpnum, title, priority, status, mesg])

#-------Rule Definitions------
#----Rule 08000 - 11999: Panorama Specific Rules
def BP08000(ip, apikey):
	bpnum = "BP08000"
	title = " Panorama Use Descriptive Rule Names"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Device Group Rules
	xpath = "/config/devices/entry[@name='localhost.localdomain']/device-group"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	
	# Query for Shared Rules
	xpath2 = "/config/shared"
	rulequery2 = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath2}
	rrule2 = requests.get('https://' + ip + '/api', params = rulequery2, verify=False)

	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall("./result/device-group/entry"):
		devgrp = entryElement.attrib['name']
		for secruleElement in entryElement.findall('pre-rulebase/security/rules/entry'):
			rule = secruleElement.attrib['name']
			if re.match("pre-rule\d{1,2}", rule):
				status = "Fail"
				mesg = "Pre-Rule %s in %s is using a default rule name" % (rule, devgrp)
				ws.append([ip, bpnum, title, priority, status, mesg])
		for secruleElement in entryElement.findall('post-rulebase/security/rules/entry'):
			rule = secruleElement.attrib['name']
			if re.match("post-rule\d", rule):
				status = "Fail"
				mesg = "Post-Rule %s in %s is using a default rule name" % (rule, devgrp)
				ws.append([ip, bpnum, title, priority, status, mesg])
	
	responseElement2 = ET.fromstring(rrule2.text)
	for entryElement in responseElement2.findall("./result/shared/pre-rulebase/security/rules/entry"):
		rule = entryElement.attrib['name']
		if re.match("shared-pre-rule\d", rule):
			status = "Fail"
			mesg = "Shared Pre-Rule %s is using a default rule name" % rule
			ws.append([ip, bpnum, title, priority, status, mesg])
	for entryElement in responseElement2.findall("./result/shared/post-rulebase/security/rules/entry"):
		rule = entryElement.attrib['name']
		if re.match("shared-post-rule\d", rule):
			status = "Fail"
			mesg = "Shared Post-Rule %s is using a default rule name" % rule
			ws.append([ip, bpnum, title, priority, status, mesg])
			
def BPPanorama(ip, apikey):
	BP01000(ip, apikey)
	BP01001(ip, apikey)
	BP01002(ip, apikey)
	BP01003(ip, apikey)
	BP01005(ip, apikey)
	BP01006(ip, apikey)
	BP01007(ip, apikey)
	BP01008(ip, apikey)
	BP01009(ip, apikey)
	BP01010(ip, apikey)
	BP01011(ip, apikey)
	BP01015(ip, apikey)
	BP01017(ip, apikey)
	BP08000(ip, apikey)

def BPUmgPan(ip, apikey):
	BP01000(ip, apikey)
	BP01001(ip, apikey)
	BP01002(ip, apikey)
	BP01003(ip, apikey)
	BP01004(ip, apikey)
	BP01005(ip, apikey)
	BP01006(ip, apikey)
	BP01007(ip, apikey)
	BP01008(ip, apikey)
	BP01009(ip, apikey)
	BP01010(ip, apikey)
	BP01011(ip, apikey)
	BP01012(ip, apikey)
	BP01013(ip, apikey)
	BP01014(ip, apikey)
	BP01015(ip, apikey)
	BP01016(ip, apikey)
	BP04000(ip, apikey)

def BPMGPan(ip, apikey):
	BP01000(ip, apikey)
	BP01001(ip, apikey)
	BP01002(ip, apikey)
	BP01003(ip, apikey)
	BP01004(ip, apikey)
	BP01005(ip, apikey)
	BP01006(ip, apikey)
	BP01007(ip, apikey)
	BP01008(ip, apikey)
	BP01009(ip, apikey)
	BP01010(ip, apikey)
	BP01011(ip, apikey)
	BP01012(ip, apikey)
	BP01013(ip, apikey)
	BP01014(ip, apikey)
	BP01015(ip, apikey)
	BP01016(ip, apikey)
	BP04000(ip, apikey)



	
def	BPTool():
	with open(fwinfo, 'r+') as csvfile:
		reader = csv.DictReader(csvfile)
		for row in reader:
			ip = row["ip_address"]
			keygen = {'type': 'keygen', 'user': row["username"], 'password': row["password"]}
			rkey = requests.post('https://' + row["ip_address"] + '/api', params=keygen, verify=False)
			rresp = ET.fromstring(rkey.content)
			apikeysearch = rresp.find('result/key')
			apikey = apikeysearch.text
			print "Generating API Key for %s" % ip
					
			#lookup device type and invoke aggregating function
			devicetype = row["device_type"]
			if devicetype == "Panorama":
				BPPanorama(ip, apikey)
			elif devicetype == "Managed-PAN":
				BPMGPan(ip, apikey)
			elif devicetype == "Unmanaged-PAN":
				BPUmgPan(ip, apikey)
			else:
				print "Device Type Not Set. Should be Panorama, Managed-PAN, or Unmanaged-PAN"

print "##############################################################"
print "#### Palo Alto Best Practices Analysis Tool               ####"
print "#### INTERNAL ONLY DO NOT DISTRIBUTE                      ####"
print "#### Version: 0.4a ALPHA (Belfast)                        ####"
print "####                                                      ####"
print "#### Written By: Jessica Ferguson                         ####"
print "#### jferguson@paloaltonetworks.com                       ####"
print "#### jferguson@infiniters.com                             ####"
print "##############################################################"
print ""
print "This software should be considered alpha code and should not be used in a production environment"
print "Here there be Dragons!!! Proceed at your Own Risk"
print ""

proceed = raw_input("Enter (y)es or (n)o: ") 
if proceed == "yes" or proceed == "y": 
	BPTool()
elif proceed == "no" or proceed == "n": 
	quit()				
			
wb.save('bp-results ' + str(datetime) + '.xlsx') 
print ""
print "##############################################################"
print "Thank you for using the Palo Alto Best Practices Analysis Tool." 
print "Your output file should be in the directory in which you launched the tool."
print "May the force be with you."
