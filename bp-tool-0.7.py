#!/usr/bin/python

import elementtree.ElementTree as ET
from sys import argv
import os
import csv
import requests
import re
from openpyxl.workbook import Workbook
from openpyxl.worksheet import filters
from openpyxl.styles import Font, Fill, Color, PatternFill
from openpyxl.formatting import FormulaRule, CellIsRule
import time
import datetime
from datetime import date

script, fwinfo = argv
now = datetime.datetime.now()
curdate = now.strftime("%Y-%m-%d")
repdate = now.strftime("%Y-%m-%d %H%M")
requests.packages.urllib3.disable_warnings()

wb = Workbook()
ws = wb.active
#ws.freeze_panes = ws.cell('A3') 

ws['A2'] = 'IP Address/FQDN'
ws['B2'] = 'BP Number'
ws['C2'] = 'Title'
ws['D2'] = 'Priority'
ws['E2'] = 'Status'
ws['F2'] = 'Description'
ws.column_dimensions["A"].width = 22.0
ws.column_dimensions["B"].width = 13.0
ws.column_dimensions["C"].width = 112.0
ws.column_dimensions["E"].width = 13.0
ws.column_dimensions["F"].width = 85.0
a1 = ws['A1']
a1.font = Font(size=20)
redFill = PatternFill(start_color='FF4500', end_color='FF4500', fill_type='solid')
greenFill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')


#-------Rule Definitions------
#----Rule 01000 - 03999: Global Rules
def BP01000(ip, apikey):
	bpnum = "BP01000"
	title = "Configure Hostname on System"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	rxpath = "result/hostname"
	xpath = "/config/devices/entry[@name='localhost.localdomain']/deviceconfig/system/hostname"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	rresp = ET.fromstring(rrule.content)
	result = rresp.find(rxpath)
	if re.match("Panorama", result.text):
		status = "Fail"
		mesg = "%s Default Panorama Hostname" % ip
	elif re.match("PA-\d{3,4}", result.text):
		status = "Fail"
		mesg = "%s Default PAN Firewall Hostname" % ip
	elif re.match("PA-VM", result.text):
		status = "Fail"
		mesg = "%s Default PAN Firewall Hostname" % ip
	else:
		status = "Pass"
		mesg = "Hostnames are Defined"
	ws.append([ip, bpnum, title, priority, status, mesg])
	
	time.sleep(sleeptime)
	
def BP01001(ip, apikey):
	bpnum = "BP01001"
	title = "Configure Domain Name on System"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	rxpath = "result/domain"
	xpath = "/config/devices/entry[@name='localhost.localdomain']/deviceconfig/system/domain"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	rresp = ET.fromstring(rrule.content)
	result = rresp.find(rxpath)
	if (result is None):
		status = "Fail"
		mesg = "%s Domain Name not Defined" % ip
	elif re.match(".+", result.text):
		status = "Pass"
		mesg = "%s Domain Name Defined" % ip
	else:
		print "FAIL!!!"
		
	ws.append([ip, bpnum, title, priority, status, mesg])
	
	time.sleep(sleeptime)
	
def BP01002(ip, apikey):
	bpnum = "BP01002"
	title = "Replace default administrative super-user account by creating a new superuser account and then deleting the admin account."
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	xpath = "/config/mgt-config/users"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	
	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall("./result/users/entry"):
		username = entryElement.attrib['name']
		for rolebasedElement in entryElement.findall('permissions/role-based'):
			superuserElement = rolebasedElement.find('superuser')
			if not superuserElement is None and superuserElement.text.lower() == 'yes' and username == 'admin':
				status = "Fail"
				mesg = "User Admin is Present"
				ws.append([ip, bpnum, title, priority, status, mesg])
			if not superuserElement is None and superuserElement.text.lower() == 'yes' and username != 'admin':
				status = "Pass"
				mesg = username + ' is superuser'
				ws.append([ip, bpnum, title, priority, status, mesg])
			superreaderElement = rolebasedElement.find('superreader')
			if not superreaderElement is None and superreaderElement.text.lower() == 'yes':
				status = "Informational"
				mesg = username + ' is superreader'
				ws.append([ip, bpnum, title, priority, status, mesg])
	
	time.sleep(sleeptime)
	
def BP01003(ip, apikey):
	bpnum = "BP01003"
	title = "Require an idle timeout value of 10 minutes for device management."
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	rxpath = "result/idle-timeout"
	xpath = "/config/devices/entry[@name='localhost.localdomain']/deviceconfig/setting/management/idle-timeout"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	rresp = ET.fromstring(rrule.content)
	result = rresp.find(rxpath)
	if result is None:
		status = "Fail"
		mesg = "Idle Timeout should be set to 10 minutes"
	elif result.text != "10":
		status = "Fail"
		mesg = "Idle Timeout should be set to 10 minutes"
	elif result.text == "10":
		status = "Pass"
		mesg = "Idle Timeout is set to 10 minutes"
	ws.append([ip, bpnum, title, priority, status, mesg])
	
	time.sleep(sleeptime)
	
def BP01004(ip, apikey):
	bpnum = "BP01004"
	title = "Configure Local User Authentication Profile"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Authentication Profile
	xpath = "/config/shared/authentication-profile"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	# Query for User Information
	xpath2 = "/config/mgt-config/users"
	rulequery2 = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath2}
	rrule2 = requests.get('https://' + ip + '/api', params = rulequery2, verify=False)
	
	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall("./result/authentication-profile/entry"):
		authprof = entryElement.attrib['name']
		for methodElement in entryElement.findall('method'):
			localdbElement = methodElement.find('local-database')
			for lockoutElement in entryElement.findall('lockout'):
				timeElement = lockoutElement.find('lockout-time')
				failElement = lockoutElement.find('failed-attempts')
				goodauth = []
				if not localdbElement is None and timeElement.text.lower() == '15' and failElement.text.lower() == '3':
					goodauth.append(authprof)
					for a in goodauth:
						status = "Pass"
						mesg = a + ' is configured correctly'
						ws.append([ip, bpnum, title, priority, status, mesg])

						responseElement2 = ET.fromstring(rrule2.text)
						usrgrp = {}
						for userElement in responseElement2.findall("./result/users/entry"):
							username = userElement.attrib['name']
							authprofElement = userElement.find('authentication-profile')
							if authprofElement is None:
								status = "Informational"
								mesg = username + ' does not have an Authentication Group assigned'
								ws.append([ip, bpnum, title, priority, status, mesg])										
							if not authprofElement is None:
								usrgrp[username] = authprofElement.text
								for username, authprofElement in usrgrp.iteritems():
									if authprofElement in goodauth:
										status = "Pass"
										mesg = username + ' is using Authentication Group ' + authprofElement
										ws.append([ip, bpnum, title, priority, status, mesg])
									
	time.sleep(sleeptime)
	
def BP01005(ip, apikey):
	bpnum = "BP01005"
	title = "Configure the firewall to verify the identity of the Palo Alto update server on the device setup page."
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	rxpath = "result/system/server-verification"
	xpath = "/config/devices/entry[@name='localhost.localdomain']/deviceconfig/system"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	rresp = ET.fromstring(rrule.content)
	result = rresp.find(rxpath)
	if result is None:
		status = "Fail"
		mesg = "Server Verification Should be Enabled"
		ws.append([ip, bpnum, title, priority, status, mesg])
	elif result.text == "yes":
		status = "Pass"
		mesg = "Server Verification is Enabled"	
		ws.append([ip, bpnum, title, priority, status, mesg])	
	elif result.text == "no":
		status = "Fail"
		mesg = "Server Verification Should be Enabled"
		ws.append([ip, bpnum, title, priority, status, mesg])
	
	time.sleep(sleeptime)
	
def BP01006(ip, apikey):
	bpnum = "BP01006"
	title = "Configure login banner."
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	rxpath = "result/system/login-banner"
	xpath = "/config/devices/entry[@name='localhost.localdomain']/deviceconfig/system"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	rresp = ET.fromstring(rrule.content)
	result = rresp.find(rxpath)
	if (result is None):
		status = "Fail"
		mesg = "Login Banner Should Be Configured"
	elif (result is not None):
		status = "Pass"
		mesg = "Login Banner Is Configured"	
		
	ws.append([ip, bpnum, title, priority, status, mesg])	
	
	time.sleep(sleeptime)


def BP01007(ip, apikey):
	bpnum = "BP01007"
	title = "Configure Geo-Location Longitude and Latitude"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	rxpath = "result/system/geo-location"
	xpath = "/config/devices/entry[@name='localhost.localdomain']/deviceconfig/system"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	rresp = ET.fromstring(rrule.content)
	result = rresp.find(rxpath)
	if (result is None):
		status = "Fail"
		mesg = "Geolocation Coordinates should be configured"
	elif (result is not None):
		status = "Pass"
		mesg = "Geolocation Coordinates are configured"	
		
	ws.append([ip, bpnum, title, priority, status, mesg])

	time.sleep(sleeptime)	
	

def BP01008(ip, apikey):
	bpnum = "BP01008"
	title = "Configure the firewall to use redundant DNS Servers"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	xpath = "/config/devices/entry[@name='localhost.localdomain']/deviceconfig/system"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)

	responseElement = ET.fromstring(rrule.text)
	for srvElement in responseElement.findall("./result/system/dns-setting/servers"):
		PrimarySrvElement = srvElement.find('primary')
		SecondarySrvElement = srvElement.find('secondary')
		if PrimarySrvElement is None: 
			status = "Fail"
			mesg = "DNS Servers are not configured"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif not PrimarySrvElement is None and SecondarySrvElement is None:
			status = "Fail"
			mesg = "Only Primary DNS Server Configured. Primary Server IP: %s" % PrimarySrvElement.text
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif not PrimarySrvElement is None and not SecondarySrvElement is None:
			status = "Pass"
			mesg = "Primary and Secondary DNS Servers Configured. Primary Server IP: %s Secondary Server IP: %s" % (PrimarySrvElement.text, SecondarySrvElement.text)
			ws.append([ip, bpnum, title, priority, status, mesg])
			
	time.sleep(sleeptime)
			
def BP01009(ip, apikey):
	bpnum = "BP01009"
	title = "Configure the firewall to use redundant NTP Servers"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	xpath = "/config/devices/entry[@name='localhost.localdomain']/deviceconfig/system"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	
	responseElement = ET.fromstring(rrule.text)
	for psrvElement in responseElement.findall("./result/system/ntp-servers/primary-ntp-server"):
		PrimarySrvElement = psrvElement.find('ntp-server-address')
		for noneElement in psrvElement.findall('authentication-type/none'):
			status = "Informational"
			mesg = "NTP Server %s configured with no authentication" % PrimarySrvElement.text
			ws.append([ip, bpnum, title, priority, status, mesg])
		for authElement in psrvElement.findall('authentication-type/symmetric-key/algorithm/*'):
			if authElement.tag == 'md5':
				status = "Fail"
				mesg = "Authencation type md5 found on NTP Server %s" % PrimarySrvElement.text
				ws.append([ip, bpnum, title, priority, status, mesg])
			if authElement.tag == 'sha1':
				status = "Pass"
				mesg = "Authencation type sha1 found on NTP Server %s" % PrimarySrvElement.text
				ws.append([ip, bpnum, title, priority, status, mesg])
				
	for ssrvElement in responseElement.findall("./result/system/ntp-servers/secondary-ntp-server"):
		SecondarySrvElement = ssrvElement.find('ntp-server-address')
		for noneElement in ssrvElement.findall('authentication-type/none'):
			status = "Informational"
			mesg = "NTP Server %s configured with no authentication" % SecondarySrvElement.text
			ws.append([ip, bpnum, title, priority, status, mesg])
		for authElement in ssrvElement.findall('authentication-type/symmetric-key/algorithm/*'):
			if authElement.tag == 'md5':
				status = "Fail"
				mesg = "Authencation type md5 found on NTP Server %s" % SecondarySrvElement.text
				ws.append([ip, bpnum, title, priority, status, mesg])
			if authElement.tag == 'sha1':
				status = "Pass"
				mesg = "Authencation type sha1 found on NTP Server %s" % SecondarySrvElement.text
				ws.append([ip, bpnum, title, priority, status, mesg])	
				
	time.sleep(sleeptime)
				
def BP01010(ip, apikey):
	bpnum = "BP01010"
	title = "Limit management interface traffic to ping and secure protocols only"
	priority = "High"
	print "Running Rule %s - %s" % (bpnum, title)
	xpath = "/config/devices/entry[@name='localhost.localdomain']/deviceconfig/system"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	
	responseElement = ET.fromstring(rrule.text)
	for srvElement in responseElement.findall("./result/system/service"):
		ssh = srvElement.find('disable-ssh')
		telnet = srvElement.find('disable-telnet')
		http = srvElement.find('disable-http')
		https = srvElement.find('disable-https')
		httpocsp = srvElement.find('disable-http-ocsp')
		snmp = srvElement.find('disable-snmp')
		icmp = srvElement.find('disable-icmp')
		uid = srvElement.find('disable-userid-service')
		uidlogssl = srvElement.find('disable-userid-syslog-listener-ssl')
		uidlogudp = srvElement.find('disable-userid-syslog-listener-udp')

		if ssh is None:
			status = "Pass"
			mesg = "SSH is Enabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif ssh.text == 'yes':
			status = "Fail"
			mesg = "SSH is Disabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif ssh.text == 'no':
			status = "Pass"
			mesg = "SSH is Enabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
			
		if telnet.text == 'yes':
			status = "Pass"
			mesg = "Telnet is Disabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif telnet.text == 'no':
			status = "Fail"
			mesg = "Telnet is Enabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		
		if http.text == 'yes':
			status = "Pass"
			mesg = "HTTP is Disabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif http.text == 'no':
			status = "Fail"
			mesg = "HTTP is Enabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
			
		if https is None:
			status = "Pass"
			mesg = "HTTPS is Enabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif https.text == 'yes':
			status = "Fail"
			mesg = "HTTPS is Disabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif https.text == 'no':
			status = "Pass"
			mesg = "HTTPS is Enabled"
			ws.append([ip, bpnum, title, priority, status, mesg])	
		
		if httpocsp is None:
			status = "Informational"
			mesg = "HTTP OSCP is Disabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif httpocsp.text == 'yes':
			status = "Informational"
			mesg = "HTTP OSCP is Disabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif httpocsp.text == 'no':
			status = "Informational"
			mesg = "HTTP OSCP is Enabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
			
		if snmp is None:
			status = "Informational"
			mesg = "SNMP Polling is Disabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif snmp.text == 'yes':
			status = "Informational"
			mesg = "SNMP Polling is Disabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif snmp.text == 'no':
			status = "Informational"
			mesg = "SNMP Polling is Enabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
			
		if icmp is None:
			status = "Informational"
			mesg = "ICMP is Disabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif icmp.text == 'yes':
			status = "Informational"
			mesg = "ICMP is Disabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif icmp.text == 'no':
			status = "Informational"
			mesg = "ICMP is Enabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
			
		if uid is None:
			status = "Informational"
			mesg = "User-ID Polling is Disabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif uid.text == 'yes':
			status = "Informational"
			mesg = "User-ID Polling is Disabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif uid.text == 'no':
			status = "Informational"
			mesg = "User-ID Polling is Enabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		
		if uidlogssl is None:
			status = "Informational"
			mesg = "User SSL Log Polling is Disabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif uidlogssl.text == 'yes':
			status = "Informational"
			mesg = "User SSL Log Polling is Disabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif uidlogssl.text == 'no':
			status = "Informational"
			mesg = "User SSL Log Polling is Enabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
			
		if uidlogudp is None:
			status = "Informational"
			mesg = "User UDP Log Polling is Disabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif uidlogudp.text == 'yes':
			status = "Informational"
			mesg = "User UDP Log Polling is Disabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif uidlogudp.text == 'no':
			status = "Informational"
			mesg = "User UDP Log Polling is Enabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
	
	time.sleep(sleeptime)

def BP01011(ip, apikey):
	bpnum = "BP01011"
	title = "Limit permitted IP Addresses to those necessary for device management"
	priority = "High"
	print "Running Rule %s - %s" % (bpnum, title)
	xpath = "/config/devices/entry[@name='localhost.localdomain']/deviceconfig/system"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	
	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall("./result/system"):
		entry = entryElement.find('permitted-ip')
		if entry is None:
			status = "Fail"
			mesg = "No Management ACL has been defined."
			ws.append([ip, bpnum, title, priority, status, mesg])
	for permElement in responseElement.findall("./result/system/permitted-ip/entry"):
		mgmtip = permElement.attrib['name']
		if mgmtip == '0.0.0.0/0':
			status = "Fail"
			mesg = "Management IP ACL has Default Entry 0.0.0.0/0"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif mgmtip:
			status = "Pass"
			mesg = "Valid Management ACL has been defined for IP %s." % mgmtip
			ws.append([ip, bpnum, title, priority, status, mesg])
			
	time.sleep(sleeptime)
			
def BP01012(ip, apikey):
	bpnum = "BP01012"
	title = "Forbid the use of password profiles."
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	xpath = "/config/mgt-config/users"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	rresp = ET.fromstring(rrule.content)
		
	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall('./result/users/entry'):
		username = entryElement.attrib['name']
		passprofile = entryElement.find('password-profile')
		if passprofile is None:
			continue
		elif passprofile.text:
			status = "Fail"
			mesg = "Password Profile %s is configured for user %s" % (username, passprofile.text)	
			ws.append([ip, bpnum, title, priority, status, mesg])
			
	time.sleep(sleeptime)
	
def BP01013(ip, apikey):
	bpnum = "BP01013"
	title = "Adjust Max Rows in CSV and User Activity Reports to 1048576"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	xpath = "/config/devices/entry[@name='localhost.localdomain']/deviceconfig/setting"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	rresp = ET.fromstring(rrule.content)
		
	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall('./result/setting/management'):
		csvExport = entryElement.find('max-rows-in-csv-export')
		pdfReport = entryElement.find('max-rows-in-pdf-report')
		if csvExport is None:
			status = "Fail"
			mesg = "Max CSV export rows is set to default"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif csvExport.text == '1048576':
			status = "Pass"
			mesg = "Max CSV export rows is set to %s" % csvExport.text
			ws.append([ip, bpnum, title, priority, status, mesg])
		else:
			status = "Fail"
			mesg = "Max CSV export rows is set to %s" % csvExport.text
			ws.append([ip, bpnum, title, priority, status, mesg])
			
		if pdfReport is None:
			status = "Fail"
			mesg = "Max PDF Report rows is set to default"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif pdfReport.text == '1048576':
			status = "Pass"
			mesg = "Max PDF Report rows is set to %s" % pdfReport.text
			ws.append([ip, bpnum, title, priority, status, mesg])
		else:
			status = "Fail"
			mesg = "Max PDF Report rows is set to %s" % pdfReport.text
			ws.append([ip, bpnum, title, priority, status, mesg])
			
	time.sleep(sleeptime)

def BP01015(ip, apikey):
	bpnum = "BP01015"
	title = "Configure minimum password complexity profile "
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Management Profile
	xpath = "/config/mgt-config/password-complexity"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	
	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall("./result"):
		passcomp = entryElement.find('password-complexity')
		if passcomp is None:
			status = "Fail"
			mesg = "Password Complexity is not Enabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		for passcompElement in entryElement.findall('password-complexity'):
			enabled = passcompElement.find('enabled')
			minlen = passcompElement.find('minimum-length')
			minup = passcompElement.find('minimum-uppercase-letters')
			minlow = passcompElement.find('minimum-lowercase-letters')
			minnum = passcompElement.find('minimum-numeric-letters')
			minspe = passcompElement.find('minimum-special-characters')
			blkusr = passcompElement.find('block-username-inclusion')
			chrdif = passcompElement.find('new-password-differs-by-characters')
			passchg = passcompElement.find('password-change-on-first-login')
			passhis = passcompElement.find('password-history-count')
			expper = passcompElement.find('password-change/expiration-period')
			expwrn = passcompElement.find('password-change/expiration-warning-period')
			expgrc = passcompElement.find('password-change/post-expiration-grace-period')
			
			# Check if Password Complexity is Enabled
			if enabled.text == 'yes':
				status = "Pass"
				mesg = "Password Complexity is Enabled"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif enabled.text == 'no':
				status = "Fail"
				mesg = "Password Complexity is not Enabled"
				ws.append([ip, bpnum, title, priority, status, mesg])

			#Check if Password Minimum Length is 12 or greater
			if minlen is None:
				status = "Fail"
				mesg = "Password Minimum Length is not Configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif minlen.text >= '12':
				status = "Pass"
				mesg = "Password Minimum Length is 12 or greater"
				ws.append([ip, bpnum, title, priority, status, mesg])
			else:
				status = "Fail"
				mesg = "Password Minimum Length is not 12 or greater"
				ws.append([ip, bpnum, title, priority, status, mesg])
			
			#Check if Password Minimum Uppercase Characters is 1 or greater
			if minup is None:
				status = "Fail"
				mesg = "Password Minimum Uppercase Characters is not Configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif minup.text >= '1':
				status = "Pass"
				mesg = "Password Minimum Uppercase Characters is 1 or greater"
				ws.append([ip, bpnum, title, priority, status, mesg])
			else:
				status = "Fail"
				mesg = "Password Minimum Uppercase Characters is not 1 or greater"
				ws.append([ip, bpnum, title, priority, status, mesg])
			
			#Check if Password Minimum Lowercase Characters is 1 or greater
			if minlow is None:
				status = "Fail"
				mesg = "Password Minimum Lowercase Characters is not Configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif minlow.text >= '1':
				status = "Pass"
				mesg = "Password Minimum Lowercase Characters is 1 or greater"
				ws.append([ip, bpnum, title, priority, status, mesg])
			else:
				status = "Fail"
				mesg = "Password Minimum Lowercase Characters is not 1 or greater"
				ws.append([ip, bpnum, title, priority, status, mesg])

			#Check if Password Minimum Numeric Characters is 1 or greater
			if minnum is None:
				status = "Fail"
				mesg = "Password Minimum Numeric Characters is not Configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif minnum.text >= '1':
				status = "Pass"
				mesg = "Password Minimum Numeric Characters is 1 or greater"
				ws.append([ip, bpnum, title, priority, status, mesg])
			else:
				status = "Fail"
				mesg = "Password Minimum Numeric Characters is not 1 or greater"
				ws.append([ip, bpnum, title, priority, status, mesg])
				
			#Check if Password Minimum Special Characters is 1 or greater
			if minspe is None:
				status = "Fail"
				mesg = "Password Minimum Special Characters is not Configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif minspe.text >= '1':
				status = "Pass"
				mesg = "Password Minimum Special Characters is 1 or greater"
				ws.append([ip, bpnum, title, priority, status, mesg])
			else:
				status = "Fail"
				mesg = "Password Minimum Special Characters is not 1 or greater"
				ws.append([ip, bpnum, title, priority, status, mesg])
				
			#Check if Block Username Inclusion is  Enabled
			if blkusr is None:
				status = "Fail"
				mesg = "Block Username Inclusion is not Configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif blkusr.text == 'yes':
				status = "Pass"
				mesg = "Block Username Inclusion is Configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			else:
				status = "Fail"
				mesg = "Block Username Inclusion is not Configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
				
			#Check if New Password Differs by Three or More Characters
			if chrdif is None:
				status = "Fail"
				mesg = "Password Differentiation is not Configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif chrdif.text >= '3':
				status = "Pass"
				mesg = "Password Differentiation is configured by 3 characters or greater"
				ws.append([ip, bpnum, title, priority, status, mesg])
			else:
				status = "Fail"
				mesg = "Password Differentiation is not Configured"
				ws.append([ip, bpnum, title, priority, status, mesg])

			#Check if Force Password Change at First Login is Enabled
			if passchg is None:
				status = "Fail"
				mesg = "Force Password Change at First Login is not Configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif passchg.text == 'yes':
				status = "Pass"
				mesg = "Force Password Change at First Login is Configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			else:
				status = "Fail"
				mesg = "Force Password Change at First Login is not Configured"
				ws.append([ip, bpnum, title, priority, status, mesg])

			#Check if Password History of 24 or Greater is Enforced
			if passhis is None:
				status = "Fail"
				mesg = "Password History of 24 or greater is not configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif passhis.text >= '24':
				status = "Pass"
				mesg = "Password History of 24 or greater is not configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			else:
				status = "Fail"
				mesg = "Password History of 24 or greater is not configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
				
			#Check if Maximum of 90 Day Password Expiration is configured
			if expper is None:
				status = "Fail"
				mesg = "Password Expiration is not configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif expper.text <= '90':
				status = "Pass"
				mesg = "Maximum of 90 Day Password Expiration is configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			else:
				status = "Fail"
				mesg = "Maximum of 90 Day Password Expiration is not configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			
			#Check if Maximum of 7 Day Expiration Warning Period is configured
			if expwrn is None:
				status = "Fail"
				mesg = "Password Expiration Warning Period is not configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif expwrn.text <= '7':
				status = "Pass"
				mesg = "Maximum of 7 Day Password Expiration Warning Period is configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			else:
				status = "Fail"
				mesg = "Maximum of 7 Day Password Expiration Warning Period is not configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
				
			#Check if Maximum of 7 Day Post Expiration Grace Period is configured
			if expgrc is None:
				status = "Fail"
				mesg = "Post Expiration Grace Period is not configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif expgrc.text <= '7':
				status = "Pass"
				mesg = "Maximum of 7 Day Post Expiration Grace Period is configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
			else:
				status = "Fail"
				mesg = "Maximum of 7 Day Post Expiration Grace Period is not configured"
				ws.append([ip, bpnum, title, priority, status, mesg])
				
	time.sleep(sleeptime)

def BP01016(ip, apikey):
	bpnum = "BP01016"
	title = "Ensure that Firewall Dynamic Updates are Processing"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for License Info
	xpath = "<request><license><info></info></license></request>"
	rulequery = {'type': 'op', 'action': 'get', 'key': apikey, 'cmd': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	responseElement = ET.fromstring(rrule.text)
	# Query for System Info
	xpath2 = "<show><system><info></info></system></show>"
	rulequery2 = {'type': 'op', 'action': 'get', 'key': apikey, 'cmd': xpath2}
	rrule2 = requests.get('https://' + ip + '/api', params = rulequery2, verify=False)
	responseElement2 = ET.fromstring(rrule2.text)
	
	for appElement in responseElement2.findall("./result/system"):
		appversion = appElement.find('app-version')
		apprelease = appElement.find('app-release-date')
		if apprelease.text =="unknown":
			status = "Fail"
			mesg = "Application Signature has never been updated"
			ws.append([ip, bpnum, title, priority, status, mesg])
		else:
			appdate = apprelease.text.split(' ')
			appdate = appdate[0].split('/')
			appdate = '-'.join(appdate)
			appdate = datetime.datetime.strptime(appdate , "%Y-%m-%d")
			testdate = datetime.datetime.strptime(curdate , "%Y-%m-%d")
			appdiff = (testdate - appdate).days
			if appdiff > 1:
				status = "Fail"
				mesg = "Application Signature Version " + appversion.text + " has not been updated in %s days" % appdiff
				ws.append([ip, bpnum, title, priority, status, mesg])
			else:
				status = "Pass"
				mesg = "Application Signature Version " + appversion.text + " is up to date"
				ws.append([ip, bpnum, title, priority, status, mesg])
	
	for entryElement in responseElement.findall("./result/licenses/entry"):
		licfeature = entryElement.find('feature')
		licexpired = entryElement.find('expired')
		activelic = []
		expiredlic = []
		if licexpired.text == "no":
			activelic.append(licfeature.text)
			for a in activelic:
				status = "Informational"
				mesg = a + " is licensed on " + ip
				ws.append([ip, bpnum, title, priority, status, mesg])
		if licexpired.text == "yes":
			expiredlic.append(licfeature.text)
			for x in expiredlic:
				status = "Informational"
				mesg = x + " license is expired on " + ip
				ws.append([ip, bpnum, title, priority, status, mesg])

		for verElement in responseElement2.findall("./result/system"):
			avversion = verElement.find('av-version')
			avrelease = verElement.find('av-release-date')
			thrversion = verElement.find('threat-version')
			thrrelease = verElement.find('threat-release-date')
			wfversion = verElement.find('wildfire-version')
			wfrelease = verElement.find('wildfire-release-date')
			urlversion = verElement.find('url-filtering-version')
			if "Threat Prevention" in activelic:
				if avrelease.text =="unknown":
					status = "Fail"
					mesg = "Anti-Virus Signature has never been updated"
					ws.append([ip, bpnum, title, priority, status, mesg])
				else:
					avdate = avrelease.text.split(' ')
					avdate = avdate[0].split('/')
					avdate = '-'.join(avdate)
					avdate = datetime.datetime.strptime(avdate , "%Y-%m-%d")
					testdate = datetime.datetime.strptime(curdate , "%Y-%m-%d")
					avdiff = (testdate - avdate).days
					if avdiff > 0:
						status = "Fail"
						mesg = "Anti-Virus Signature Version " + avversion.text + " has not been updated in %s days" % avdiff
						ws.append([ip, bpnum, title, priority, status, mesg])
					else:
						status = "Pass"
						mesg = "Anti-Virus Signature Version " + avversion.text + " is up to date"
						ws.append([ip, bpnum, title, priority, status, mesg])
				
				if thrrelease.text =="unknown":
					status = "Fail"
					mesg = "Threat Signature has never been updated"
					ws.append([ip, bpnum, title, priority, status, mesg])
				else:
					thrdate = thrrelease.text.split(' ')
					thrdate = thrdate[0].split('/')
					thrdate = '-'.join(thrdate)
					thrdate = datetime.datetime.strptime(thrdate , "%Y-%m-%d")
					testdate = datetime.datetime.strptime(curdate , "%Y-%m-%d")
					thrdiff = (testdate - thrdate).days
					if thrdiff > 1:
						status = "Fail"
						mesg = "Threat Signature Version " + thrversion.text + " has not been updated in %s days" % thrdiff
						ws.append([ip, bpnum, title, priority, status, mesg])
					else:
						status = "Pass"
						mesg = "Threat Signature Version " + thrversion.text + " is up to date"
						ws.append([ip, bpnum, title, priority, status, mesg])
					
			if "WildFire License" in activelic:
				if wfrelease.text =="unknown":
					status = "Fail"
					mesg = "WildFire Signature has never been updated"
					ws.append([ip, bpnum, title, priority, status, mesg])
				else:
					wfdate = wfrelease.text.split(' ')
					wfdate = wfdate[0].split('/')
					wfdate = '-'.join(wfdate)
					wfdate = datetime.datetime.strptime(wfdate , "%Y-%m-%d")
					testdate = datetime.datetime.strptime(curdate , "%Y-%m-%d")
					wfdiff = (testdate - wfdate).days
					if wfdiff > 0:
						status = "Fail"
						mesg = "WildFire Signature Version " + wfversion.text + " has not been updated in %s days" % wfdiff
						ws.append([ip, bpnum, title, priority, status, mesg])
					else:
						status = "Pass"
						mesg = "WildFire Signature Version " + wfversion.text + " is up to date"
						ws.append([ip, bpnum, title, priority, status, mesg])
						
	time.sleep(sleeptime)

def BP01017(ip, apikey):
	bpnum = "BP01017"
	title = "Ensure that Dynamic Update Times are Configured Properly"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for License Info
	xpath = "/config/devices/entry[@name='localhost.localdomain']/deviceconfig/system"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	responseElement = ET.fromstring(rrule.text)
		
	for entryElement in responseElement.findall("./result/system/update-schedule"):
		thrupdate = entryElement.find('threats')
		avupdate = entryElement.find('anti-virus')
		wfupdate = entryElement.find('wildfire')

		
		if not thrupdate is None:
			for recElement in entryElement.findall('threats/recurring'):
				weekly = recElement.find('weekly')
				daily = recElement.find('daily')
				if weekly:
					status = "Fail"
					mesg = "Threat and Application Updates are configured to install weekly on %s" % ip
					ws.append([ip, bpnum, title, priority, status, mesg])
				elif daily:
					for dayElement in recElement.findall('daily'):
						timeat = dayElement.find('at')
						action = dayElement.find('action')
						if action.text == 'download-only':
							status = "Fail"
							mesg = "Threat and Application Updates are configured to download only"
							ws.append([ip, bpnum, title, priority, status, mesg])
						if action.text == 'download-and-install':
							status = "Pass"
							mesg = "Threat and Application Updates are configured to download and install daily"
							ws.append([ip, bpnum, title, priority, status, mesg])
		
		if not avupdate is None:
			for recElement in entryElement.findall('anti-virus/recurring'):
				weekly = recElement.find('weekly')
				daily = recElement.find('daily')
				hourly = recElement.find('hourly')
				if weekly:
					status = "Fail"
					mesg = "Anti-Virus Updates are configured to install weekly on %s" % ip
					ws.append([ip, bpnum, title, priority, status, mesg])
				elif daily:
					status = "Fail"
					mesg = "Anti-Virus Updates are configured to install daily on %s" % ip
					ws.append([ip, bpnum, title, priority, status, mesg])
				elif hourly:
					for hourElement in recElement.findall('hourly'):
						timeat = hourElement.find('at')
						action = hourElement.find('action')
						if action.text == 'download-only':
							status = "Fail"
							mesg = "Anti-Virus Updates are configured to download only"
							ws.append([ip, bpnum, title, priority, status, mesg])
						if action.text == 'download-and-install':
							status = "Pass"
							mesg = "Anti-Virus Updates are configured to download and install hourly"
							ws.append([ip, bpnum, title, priority, status, mesg])
		
		if not wfupdate is None:
			for recElement in entryElement.findall('wildfire/recurring'):
				fifteenmin = recElement.find('every-15-mins')
				thirtymin = recElement.find('every-30-mins')
				hourly = recElement.find('every-hour')
				if hourly:
					status = "Fail"
					mesg = "WildFire Updates are configured to install hourly on %s" % ip
					ws.append([ip, bpnum, title, priority, status, mesg])
				elif thirtymin:
					status = "Fail"
					mesg = "WildFire Updates are configured to install every half hour on %s" % ip
					ws.append([ip, bpnum, title, priority, status, mesg])
				elif fifteenmin:
					for fifElement in recElement.findall('every-15-mins'):
						timeat = fifElement.find('at')
						action = fifElement.find('action')
						if action.text == 'download-only':
							status = "Fail"
							mesg = "WildFire Updates are configured to download only"
							ws.append([ip, bpnum, title, priority, status, mesg])
						if action.text == 'download-and-install':
							status = "Pass"
							mesg = "WildFire Updates are configured to download and install every fifteen minutes"
							ws.append([ip, bpnum, title, priority, status, mesg])
							
	time.sleep(sleeptime)

def BP01018(ip, apikey):
	bpnum = "BP01018"
	title = "Validate that the default Admin password has been changed."
	priority = "Medium"
	print "Running Rule %s - %s" % (bpnum, title)
	xpath = "/config/mgt-config/users"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	
	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall("./result/users/entry"):
		username = entryElement.attrib['name']
		phashElement = entryElement.find('phash')
		if phashElement is None:
			continue
		if phashElement.text == 'fnRL/G5lXVMug':
			status = "Fail"
			mesg = "Default password 'admin' configured on account %s" % username
			ws.append([ip, bpnum, title, priority, status, mesg])
			
	time.sleep(sleeptime)
	
#-------Rule Definitions------
#----Rule 04000 - 07999: Firewall Specific Rules
def BP04000(ip, apikey):
	bpnum = "BP04000"
	title = "Firewall Should Use Descriptive Rule Names"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Management Profile
	xpath = "/config/devices/entry[@name='localhost.localdomain']/vsys"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall("./result/vsys/entry"):
		vsys = entryElement.attrib['name']
		for secruleElement in entryElement.findall('rulebase/security/rules/entry'):
			rule = secruleElement.attrib['name']
			if re.match("rule\d{1,2}", rule):
				status = "Fail"
				mesg = "%s in %s is using a default rule name" % (rule, vsys)
				ws.append([ip, bpnum, title, priority, status, mesg])
				
	time.sleep(sleeptime)

def BP04001(ip, apikey):
	bpnum = "BP04001"
	title = "Firewall eTAC Recommended Versions of Code"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Management Profile
	xpath = "<show><system><info></info></system></show>"
	rulequery = {'type': 'op', 'action': 'get', 'key': apikey, 'cmd': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	
	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall("./result/system"):
		version = entryElement.find('sw-version')
		if version.text == "6.0.8" or version.text == "6.0.9" or version.text == "6.0.10" or version.text == "6.1.4" or version.text == "6.1.5":
			status = "Pass"
			mesg = "Current Version: %s - Firewall is running an eTAC Recommended Version of Code" % version.text
			ws.append([ip, bpnum, title, priority, status, mesg])
		else:
			status = "Fail"
			mesg = "Current Version: %s - Firewall is not running an eTAC Recommended Version of Code. For more information refer to https://intranet.paloaltonetworks.com/docs/DOC-4857" % version.text
			ws.append([ip, bpnum, title, priority, status, mesg])
			
	time.sleep(sleeptime)

def BP04002(ip, apikey):	
	bpnum = "BP04002"
	title = "Firewall Should Use RADIUS for User Authentication"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Configured Authentication Profiles
	xpath = "/config/devices/entry[@name='localhost.localdomain']/deviceconfig/system"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	# Query for Authentication Profile Protocol
	xpath2 = "/config/shared"
	rulequery2 = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath2}
	rrule2 = requests.get('https://' + ip + '/api', params = rulequery2, verify=False)
		
	responseElement = ET.fromstring(rrule.text) 
	for entryElement in responseElement.findall("./result/system"):
		sysauth = entryElement.find('authentication-profile')
		sysauthprof = []
		if sysauth is None:
			status = "Fail"
			mesg = "Authentication Profile is not Configured, Using local authentication."
			ws.append([ip, bpnum, title, priority, status, mesg])
		else:
			sysauthprof.append(sysauth.text)
			status = "Informational"
			mesg = "Authentication Profile %s is configured" % sysauth.text
			ws.append([ip, bpnum, title, priority, status, mesg])
		
		responseElement2 = ET.fromstring(rrule2.text)
		for authprofElement in responseElement2.findall("./result/shared/authentication-profile"):
			for entryElement in authprofElement.findall('entry'):
				entryName = entryElement.attrib['name']
				if entryName in sysauthprof:
					for methodElement in entryElement.findall('method'):
						for methodTypeElement in methodElement:
							methodType = methodTypeElement.tag.upper()
							status = "Pass"
							mesg = "Firewall Authentication Profile %s is using %s" % (entryName , methodType)
							ws.append([ip, bpnum, title, priority, status, mesg])
	
	time.sleep(sleeptime)

def BP04003(ip, apikey):
	bpnum = "BP04003"
	title = "Firewall Alert on Invalid Subnet Masks"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Device Group Objects
	xpath = "/config/devices/entry[@name='localhost.localdomain']/vsys"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	# Query for Shared Objects
	xpath2 = "/config/shared"
	rulequery2 = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath2}
	rrule2 = requests.get('https://' + ip + '/api', params = rulequery2, verify=False)

	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall('./result/vsys/entry'):
		vsys = entryElement.attrib['name']
		## Check VSYS Rulebase for Invalid Mask in rule
		for secruleElement in entryElement.findall('rulebase/security/rules/entry'):
			rule = secruleElement.attrib['name']
			for srcaddrElement in secruleElement.findall('source'):
				srcaddr = srcaddrElement.find('member')
				if srcaddr is None:
					continue
				elif re.match("^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\/([1-7]{1})$", srcaddr.text):
					status = "Fail"
					mesg = "Invalid Subnet Mask %s found in vsys '%s' Rule '%s'" % (srcaddr.text, vsys, rule)
					ws.append([ip, bpnum, title, priority, status, mesg])
			for dstaddrElement in secruleElement.findall('destination'):
				dstaddr = dstaddrElement.find('member')
				if dstaddr is None:
					continue
				elif re.match("^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\/([1-7]{1})$", dstaddr.text):
					status = "Fail"
					mesg = "Invalid Subnet Mask %s found in vsys '%s' Rule '%s'" % (dstaddr.text, vsys, rule)
					ws.append([ip, bpnum, title, priority, status, mesg])
		
		## Check VSYS Address Objects for Invalid Mask						
		for vsysaddrElement in entryElement.findall('address/entry'):
			address = vsysaddrElement.attrib['name']
			ipmask = vsysaddrElement.find('ip-netmask')
			if ipmask is None:
				continue
			elif re.match("^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\/([1-7]{1})$", ipmask.text):
				status = "Fail"
				mesg = "Invalid Subnet Mask %s found in vsys '%s' Address Object '%s'" % (ipmask.text, vsys, address)
				ws.append([ip, bpnum, title, priority, status, mesg])
			
	
	responseElement2 = ET.fromstring(rrule2.text)
	for entryElement in responseElement2.findall('./result/shared'):
		## Check Shared Address Objects for Invalid Mask							
		for srdaddrElement in entryElement.findall('address/entry'):
			address = srdaddrElement.attrib['name']
			ipmask = srdaddrElement.find('ip-netmask')
			if ipmask is None:
				continue
			elif re.match("^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\/([1-7]{1})$", ipmask.text):
				status = "Fail"
				mesg = "Invalid Subnet Mask %s found in Shared Address Object '%s'" % (ipmask.text, address)
				ws.append([ip, bpnum, title, priority, status, mesg])
				
	time.sleep(sleeptime)

def BP04004(ip, apikey):	
	bpnum = "BP04004"
	title = "Define Syslog Servers on Firewall"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Shared Config
	xpath = "/config/shared"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)

	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall('./result/shared'):
		syslog = entryElement.find('log-settings/syslog')
		if syslog is None:
			status = "Fail"
			mesg = "Syslog not configured on system"
			ws.append([ip, bpnum, title, priority, status, mesg])
		else:
			for logsetElement in entryElement.findall('log-settings/syslog/entry'):
				profile = logsetElement.attrib['name']
				for srvElement in logsetElement.findall('server/entry'):
					server = srvElement.attrib['name']
					facility = srvElement.find('facility')
					ipaddr = srvElement.find('server')
					if facility.text == 'LOG_USER':
						status = "Pass"
						mesg = "SYSLOG Server %s is configured for IP %s using facility %s" % (server, ipaddr.text, facility.text)
						ws.append([ip, bpnum, title, priority, status, mesg])
					elif facility.text != 'LOG_USER':
						status = "Fail"
						mesg = "SYSLOG Server %s is configured for IP %s using facility %s" % (server, ipaddr.text, facility.text)
						ws.append([ip, bpnum, title, priority, status, mesg])
	
	time.sleep(sleeptime)
	
def BP04005(ip, apikey):
	bpnum = "BP04005"
	title = "Define SNMP TRAP Servers on Firewall"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Shared Config
	xpath = "/config/shared"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)

	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall('./result/shared'):
		snmpTrap = entryElement.find('log-settings/snmptrap')
		if snmpTrap is None:
			status = "Fail"
			mesg = "SNMP Trap Server not configured on system"
			ws.append([ip, bpnum, title, priority, status, mesg])
		else:
			for snmpElement in entryElement.findall('log-settings/snmptrap/entry'):
				profile = snmpElement.attrib['name']
				for v2srvElement in snmpElement.findall('version/v2c/server/entry'):
					server = v2srvElement.attrib['name']
					ipaddr = v2srvElement.find('manager')
					status = "Pass"
					mesg = "SNMP v2 Trap Server %s is configured for IP %s" % (server, ipaddr.text)
					ws.append([ip, bpnum, title, priority, status, mesg])
				for v3srvElement in snmpElement.findall('version/v3/server/entry'):
					server = v3srvElement.attrib['name']
					ipaddr = v3srvElement.find('manager')
					status = "Pass"
					mesg = "SNMP v3 Trap Server %s is configured for IP %s" % (server, ipaddr.text)
					ws.append([ip, bpnum, title, priority, status, mesg])
	
	time.sleep(sleeptime)

def BP04006(ip, apikey):
	bpnum = "BP04006"
	title = "Define NetFlow Collectors on Firewall"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Shared Config
	xpath = "/config/shared"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)

	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall('./result/shared'):
		nfProfile = entryElement.find('server-profile/netflow')
		if nfProfile is None:
			status = "Fail"
			mesg = "NetFlow Collector Server not configured on system"
			ws.append([ip, bpnum, title, priority, status, mesg])
		else:
			for nfprofElement in entryElement.findall('server-profile/netflow/entry'):
				profile = nfprofElement.attrib['name']
				refmin = nfprofElement.find('template-refresh-rate/minutes')
				refpac = nfprofElement.find('template-refresh-rate/packets')
				activetime = nfprofElement.find('active-timeout')
				panfield = nfprofElement.find('export-enterprise-fields')
				status = "Pass"
				mesg = "NetFlow Profile %s configured with Template Refresh of %s minutes or %s packets with an active timeout of %s minutes. PAN-OS Field Types is enabled: %s" % (profile, refmin.text, refpac.text, activetime.text, panfield.text)
				ws.append([ip, bpnum, title, priority, status, mesg])
				for serverElement in nfprofElement.findall('server/entry'):
					nfserv = serverElement.attrib['name']
					host = serverElement.find('host')
					port = serverElement.find('port')
					status = "Informational"
					mesg = "NetFlow Profile %s is configured with server %s %s on port %s" % (profile, nfserv, host.text, port.text)
					ws.append([ip, bpnum, title, priority, status, mesg])
					
	time.sleep(sleeptime)
	
def BP04007(ip, apikey):
	bpnum = "BP04007"
	title = "Validate That a NetFlow Profile is Assigned to an Interface"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Shared Config
	xpath = "/config/devices/entry[@name='localhost.localdomain']/network/interface"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)

	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall('./result/interface/ethernet/entry'):
		intfName = entryElement.attrib['name']
		nfprof = entryElement.find('layer3/netflow-profile')
		if not nfprof is None:
			status = "Informational"
			mesg = "NetFlow Profile %s is configured on interface %s" % (nfprof.text, intfName)
			ws.append([ip, bpnum, title, priority, status, mesg])
			
	time.sleep(sleeptime)
	
def BP04008(ip, apikey):
	bpnum = "BP04008"
	title = "Configure target e-mail account to receive critical and high severity threat, system, and Wildfire log events."
	priority = "Informational"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Shared Config
	xpath = "/config/shared"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)

	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall('./result/shared'):
		email = entryElement.find('log-settings/email/')
		if email is None:
			status = "Informational"
			mesg = "Email Profile is not configured on system"
			ws.append([ip, bpnum, title, priority, status, mesg])
		else:
			for emailsetElement in entryElement.findall('log-settings/email/entry'):
				profile = emailsetElement.attrib['name']
				for srvElement in emailsetElement.findall('server/entry'):
					server = srvElement.attrib['name']
					gateway = srvElement.find('gateway')
					to = srvElement.find('to')
					status = "Informational"
					mesg = "Email Profile %s is configured for gateway %s using email %s" % (profile, gateway.text, to.text)
					ws.append([ip, bpnum, title, priority, status, mesg])
					
	time.sleep(sleeptime)

def BP04009(ip, apikey):
	bpnum = "BP04009"
	title = "Configure Firewall System Logs to Forward to Panorama, Syslog, SNMP, or eMail."
	priority = "Medium"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Shared Config
	xpath = "/config/shared"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)

	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall('./result/shared'):
		system = entryElement.find('log-settings/system/')
		informational = entryElement.find('log-settings/system/informational')
		low = entryElement.find('log-settings/system/low')
		medium = entryElement.find('log-settings/system/medium')
		high = entryElement.find('log-settings/system/high')
		critical = entryElement.find('log-settings/system/critical')
		if system is None:
			status = "Fail"
			mesg = "System Logging is not configured."
			ws.append([ip, bpnum, title, priority, status, mesg])
			break
		
		if not informational is None:
			for infoElement in entryElement.findall('log-settings/system/informational'):
				loggrp = {}
				snmp = infoElement.find('send-snmptrap/using-snmptrap-setting')
				email = infoElement.find('send-email/using-email-setting')
				syslog = infoElement.find('send-syslog/using-syslog-setting')
				panorama = infoElement.find('send-to-panorama')
				if not snmp is None:
					loggrp['snmp'] = snmp.text
				if not email is None:
					loggrp['email'] = email.text
				if not syslog is None:
					loggrp['syslog'] = syslog.text
				if not panorama is None:
					loggrp['Panorama'] = panorama.text
				for type, profile in loggrp.iteritems():
					status = "Pass"
					mesg = "System Informational Logging configured for " + type + ": " + profile
					ws.append([ip, bpnum, title, priority, status, mesg])
		else:
			status = "Informational"
			mesg = "Informational Logging Profiles not Configured."
			ws.append([ip, bpnum, title, priority, status, mesg])

		
		if not low is None:
			for infoElement in entryElement.findall('log-settings/system/low'):
				loggrp = {}
				snmp = infoElement.find('send-snmptrap/using-snmptrap-setting')
				email = infoElement.find('send-email/using-email-setting')
				syslog = infoElement.find('send-syslog/using-syslog-setting')
				panorama = infoElement.find('send-to-panorama')
				if not snmp is None:
					loggrp['snmp'] = snmp.text
				if not email is None:
					loggrp['email'] = email.text
				if not syslog is None:
					loggrp['syslog'] = syslog.text
				if not panorama is None:
					loggrp['Panorama'] = panorama.text
				for type, profile in loggrp.iteritems():
					status = "Pass"
					mesg = "System Low Logging configured for " + type + ": " + profile
					ws.append([ip, bpnum, title, priority, status, mesg])
		else:
			status = "Informational"
			mesg = "Low Logging Profiles not Configured."
			ws.append([ip, bpnum, title, priority, status, mesg])
			
		if not medium is None:
			for infoElement in entryElement.findall('log-settings/system/medium'):
				loggrp = {}
				snmp = infoElement.find('send-snmptrap/using-snmptrap-setting')
				email = infoElement.find('send-email/using-email-setting')
				syslog = infoElement.find('send-syslog/using-syslog-setting')
				panorama = infoElement.find('send-to-panorama')
				if not snmp is None:
					loggrp['snmp'] = snmp.text
				if not email is None:
					loggrp['email'] = email.text
				if not syslog is None:
					loggrp['syslog'] = syslog.text
				if not panorama is None:
					loggrp['Panorama'] = panorama.text
				for type, profile in loggrp.iteritems():
					status = "Pass"
					mesg = "System Medium Logging configured for " + type + ": " + profile
					ws.append([ip, bpnum, title, priority, status, mesg])
		else:
			status = "Fail"
			mesg = "Medium Logging Profiles not Configured."
			ws.append([ip, bpnum, title, priority, status, mesg])
		
		if not high is None:
			for infoElement in entryElement.findall('log-settings/system/high'):
				loggrp = {}
				snmp = infoElement.find('send-snmptrap/using-snmptrap-setting')
				email = infoElement.find('send-email/using-email-setting')
				syslog = infoElement.find('send-syslog/using-syslog-setting')
				panorama = infoElement.find('send-to-panorama')
				if not snmp is None:
					loggrp['snmp'] = snmp.text
				if not email is None:
					loggrp['email'] = email.text
				if not syslog is None:
					loggrp['syslog'] = syslog.text
				if not panorama is None:
					loggrp['Panorama'] = panorama.text
				for type, profile in loggrp.iteritems():
					status = "Pass"
					mesg = "System High Logging configured for " + type + ": " + profile
					ws.append([ip, bpnum, title, priority, status, mesg])
		else:
			status = "Fail"
			mesg = "High Logging Profiles not Configured."
			ws.append([ip, bpnum, title, priority, status, mesg])
				
		if not critical is None:
			for infoElement in entryElement.findall('log-settings/system/critical'):
				loggrp = {}
				snmp = infoElement.find('send-snmptrap/using-snmptrap-setting')
				email = infoElement.find('send-email/using-email-setting')
				syslog = infoElement.find('send-syslog/using-syslog-setting')
				panorama = infoElement.find('send-to-panorama')
				if not snmp is None:
					loggrp['snmp'] = snmp.text
				if not email is None:
					loggrp['email'] = email.text
				if not syslog is None:
					loggrp['syslog'] = syslog.text
				if not panorama is None:
					loggrp['Panorama'] = panorama.text
				for type, profile in loggrp.iteritems():
					status = "Pass"
					mesg = "System Critical Logging configured for " + type + ": " + profile
					ws.append([ip, bpnum, title, priority, status, mesg])
		else:
			status = "Fail"
			mesg = "Critical Logging Profiles not Configured."
			ws.append([ip, bpnum, title, priority, status, mesg])

	time.sleep(sleeptime)

def BP04010(ip, apikey):	
	bpnum = "BP04010"
	title = "Configure Firewall Config Logs to Forward to Panorama, Syslog, SNMP, or eMail."
	priority = "Medium"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Shared Config
	xpath = "/config/shared"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)

	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall('./result/shared'):
		config = entryElement.find('log-settings/config/')
		if config is None:
			status = "Fail"
			mesg = "Configuration Change Logging is not configured."
			ws.append([ip, bpnum, title, priority, status, mesg])
			break
		
		for infoElement in entryElement.findall('log-settings/config/any'):
			loggrp = {}
			snmp = infoElement.find('send-snmptrap/using-snmptrap-setting')
			email = infoElement.find('send-email/using-email-setting')
			syslog = infoElement.find('send-syslog/using-syslog-setting')
			panorama = infoElement.find('send-to-panorama')
			if not snmp is None:
				loggrp['snmp'] = snmp.text
			if not email is None:
				loggrp['email'] = email.text
			if not syslog is None:
				loggrp['syslog'] = syslog.text
			if not panorama is None:
				loggrp['Panorama'] = panorama.text
			for type, profile in loggrp.iteritems():
				status = "Pass"
				mesg = "Configuration Change Logging configured for " + type + ": " + profile
				ws.append([ip, bpnum, title, priority, status, mesg])
	
	time.sleep(sleeptime)
	
def BP04011(ip, apikey):
	bpnum = "BP04011"
	title = "Enable Log on High DP Load"
	priority = "High"
	print "Running Rule %s - %s" % (bpnum, title)
	xpath = "/config/devices/entry[@name='localhost.localdomain']/deviceconfig/setting"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	
	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall("./result/setting/management"):
		loghighdp = entryElement.find('enable-log-high-dp-load')
		if loghighdp is None:
			status = "Fail"
			mesg = "High DP Logging Not Enabled."
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif loghighdp.text == 'yes':
			status = "Pass"
			mesg = "High DP Logging Is Enabled."
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif loghighdp.text == 'no':
			status = "Fail"
			mesg = "High DP Logging Not Enabled."
			ws.append([ip, bpnum, title, priority, status, mesg])
			
	time.sleep(sleeptime)
			
def BP04012(ip, apikey):
	bpnum = "BP04012"
	title = "Secure Production Interface Management Profiles"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Interface Management Profile
	xpath = "/config/devices/entry[@name='localhost.localdomain']/network/profiles/interface-management-profile"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	# Query for Interface Information
	xpath2 = "/config/devices/entry[@name='localhost.localdomain']/network/interface/ethernet"
	rulequery2 = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath2}
	rrule2 = requests.get('https://' + ip + '/api', params = rulequery2, verify=False)
	
	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall("./result/interface-management-profile/entry"):
		profile = entryElement.attrib['name']
		ssh = entryElement.find('ssh')
		https = entryElement.find('https')
		ping = entryElement.find('ping')
		http = entryElement.find('http')
		telnet = entryElement.find('telnet')
		goodauth = []
		badauth = []
		if http is None and telnet is None:
			goodauth.append(profile)
			for a in goodauth:
				status = "Pass"
				mesg = a + ' is configured correctly'
				ws.append([ip, bpnum, title, priority, status, mesg])
		if not http is None and http.text == 'yes' or not telnet is None and telnet.text == 'yes':
			badauth.append(profile)
			for a in badauth:
				status = "Fail"
				mesg = a + ' is configured incorrectly'
				ws.append([ip, bpnum, title, priority, status, mesg])
		if not http is None and http.text == 'no' or not telnet is None and telnet.text == 'no':
			goodauth.append(profile)
			for a in goodauth:
				status = "Pass"
				mesg = a + ' is configured correctly'
				ws.append([ip, bpnum, title, priority, status, mesg])

		responseElement2 = ET.fromstring(rrule2.text)
		for intfElement in responseElement2.findall("./result/ethernet/entry"):
			intfgrp = {}
			interface = intfElement.attrib['name']
			mgmtprofElement = intfElement.find('layer3/interface-management-profile')
			if mgmtprofElement is None:
				status = "Informational"
				mesg = interface + ' does not have an Management Profile assigned'
				ws.append([ip, bpnum, title, priority, status, mesg])										
			if not mgmtprofElement is None and not interface in intfgrp:
				intfgrp[interface] = mgmtprofElement.text
				for interface, mgmtprofElement in intfgrp.iteritems():
					if mgmtprofElement in goodauth:
						status = "Pass"
						mesg = interface + ' is using Secure Management Profile ' + mgmtprofElement
						ws.append([ip, bpnum, title, priority, status, mesg])
					if mgmtprofElement in badauth:
						status = "Fail"
						mesg = interface + ' is using unsecure Management Profile ' + mgmtprofElement
						ws.append([ip, bpnum, title, priority, status, mesg])
						
	time.sleep(sleeptime)

def BP04013(ip, apikey):
	bpnum = "BP04013"
	title = "Restrict Production Interface Management Profiles Source Addresses"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Management Profile
	xpath = "/config/devices/entry[@name='localhost.localdomain']/network/profiles/interface-management-profile"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	
	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall("./result/interface-management-profile/entry"):
		profile = entryElement.attrib['name']
		permip = entryElement.find('permitted-ip')
		if permip is None:
			status = "Fail"
			mesg = "No Management ACL has been defined for " + profile
			ws.append([ip, bpnum, title, priority, status, mesg])
		for ipElement in entryElement.findall("permitted-ip/entry"):
			mgmtip = ipElement.attrib['name']
			if mgmtip == '0.0.0.0/0':
				status = "Fail"
				mesg = "Management IP ACL has Default Entry 0.0.0.0/0 in Profile %s" % profile
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif mgmtip:
				status = "Pass"
				mesg = "Valid Management ACL has been defined for IP %s. in Profile %s" % (mgmtip, profile)
				ws.append([ip, bpnum, title, priority, status, mesg])
				
	time.sleep(sleeptime)
	
def BP04014(ip, apikey):
	bpnum = "BP04014"
	title = "Require a fully-synchronized High Availability peer"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	xpath = "<show><high-availability><state></state></high-availability></show>"
	rulequery = {'type': 'op', 'action': 'get', 'key': apikey, 'cmd': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	rresp = ET.fromstring(rrule.content)
		
	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall('./result/group'):
		configSync = entryElement.find('running-sync')
		if configSync is None:
			continue
		elif configSync.text == 'not synchronized':
			status = "Fail"
			mesg = "Running Config is Not Synced to Peer on %s" % ip
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif configSync.text == 'synchronized':
			status = "Pass"
			mesg = "Running Config is Synced to Peer on %s" % ip
			ws.append([ip, bpnum, title, priority, status, mesg])
			
	time.sleep(sleeptime)
	
def BP04015(ip, apikey):	
	bpnum = "BP04015"
	title = "For High Availability, require Link Monitoring, Path Monitoring, or both"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	xpath = "<show><high-availability><all></all></high-availability></show>"
	rulequery = {'type': 'op', 'action': 'get', 'key': apikey, 'cmd': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	rresp = ET.fromstring(rrule.content)
		
	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall('./result/group'):
		linkMonitor = entryElement.find('link-monitoring/enabled')
		pathMonitor = entryElement.find('path-monitoring/enabled')
		if linkMonitor or pathMonitor is None:
			continue
		elif linkMonitor.text == 'no' and pathMonitor.text == 'no':
			status = "Fail"
			mesg = "Path Monitoring and Link Monitoring not configured."
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif linkMonitor.text == 'yes' and pathMonitor.text == 'yes':
			status = "Pass"
			mesg = "Path Monitoring and Link Monitoring are configured."
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif linkMonitor.text == 'yes':
			status = "Pass"
			mesg = "Link Monitoring is Configured"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif pathMonitor.text == 'yes':
			status = "Pass"
			mesg = "Path Monitoring is Configured"
			ws.append([ip, bpnum, title, priority, status, mesg])
	
	time.sleep(sleeptime)
	
def BP04016(ip, apikey):
	bpnum = "BP04016"
	title = "Forbid simultaneously enabling the Preemptive option, and configuring the Passive Link State to shutdown simultaneously."
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	xpath = "<show><high-availability><all></all></high-availability></show>"
	rulequery = {'type': 'op', 'action': 'get', 'key': apikey, 'cmd': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	rresp = ET.fromstring(rrule.content)
		
	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall('./result/group'):
		localLinkState = entryElement.find('local-info/active-passive/passive-link-state')
		localPreempt = entryElement.find('local-info/preemptive')
		if localLinkState or localPreempt is None:
			continue
		elif localLinkState.text == 'shutdown' and localPreempt.text == 'yes':
			status = "Fail"
			mesg = "Link State should not be shutdown with Preemptive Enabled."
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif localLinkState.text == 'auto' and localPreempt.text == 'yes':
			status = "Pass"
			mesg = "Link State set to Auto with Preemptive Enabled."
			ws.append([ip, bpnum, title, priority, status, mesg])
			
	time.sleep(sleeptime)

def BP04017(ip, apikey):
	bpnum = "BP04017"
	title = "Require IP-to-username mapping for user traffic"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Management Profile
	xpath = "/config/devices/entry[@name='localhost.localdomain']/vsys"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	
	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall("./result/vsys/entry"):
		vsys = entryElement.attrib['name']
		uida = entryElement.find('user-id-agent')
		uidc = entryElement.find('user-id-collector')
		tsagent = entryElement.find('ts-agent')
		if uida is None and uidc is None and tsagent is None:
			status = "Informational"
			mesg = "User-ID is not configured on this firewall."
			ws.append([ip, bpnum, title, priority, status, mesg])
		if uida:
			status = "Pass"
			mesg = "User-ID Agent Server configured."
			ws.append([ip, bpnum, title, priority, status, mesg])
			for uidaElement in entryElement.findall('user-id-agent/entry'):
				profile = uidaElement.attrib['name']
				host = uidaElement.find('host')
				port = uidaElement.find('port')
				status = "Informational"
				mesg = "User-ID Agent Server %s is configured on host %s:%s" % (profile, host.text, port.text)
				ws.append([ip, bpnum, title, priority, status, mesg])
		if uidc:
			status = "Pass"
			mesg = "Internal User-ID Agent configured."
			ws.append([ip, bpnum, title, priority, status, mesg])
			for uidcElement in entryElement.findall('user-id-collector/server-monitor/entry'):
				profile = uidcElement.attrib['name']
				ad = uidcElement.find('active-directory')
				adhost = uidcElement.find('active-directory/host')
				ex = uidcElement.find('exchange')
				exhost = uidcElement.find('exchange/host')
				sys = uidcElement.find('syslog')
				syshost = uidcElement.find('syslog/address')
				edir = uidcElement.find('e-directory')
				edirprof = uidcElement.find('e-directory/server-profile')
				if not ad is None:
					status = "Informational"
					mesg = "Active Directory Profile %s is configured for server %s" % (profile, adhost.text)
					ws.append([ip, bpnum, title, priority, status, mesg])
				if not ex is None:
					status = "Informational"
					mesg = "Exchange Profile %s is configured for server %s" % (profile, exhost.text)
					ws.append([ip, bpnum, title, priority, status, mesg])
				if not sys is None:
					status = "Informational"
					mesg = "Syslog Sender %s is configured for server %s" % (profile, syshost.text)
					ws.append([ip, bpnum, title, priority, status, mesg])
				if not edir is None:
					status = "Informational"
					mesg = "Novell e-directory profile %s is configured for LDAP server profile %s" % (profile, edirprof.text)
					ws.append([ip, bpnum, title, priority, status, mesg])
			for wmiElement in entryElement.findall('user-id-collector/setting'):
				wmiAccount = wmiElement.find('wmi-account')
				status = "Informational"
				mesg = "WMI Service Account in Use is %s" % wmiAccount.text
				ws.append([ip, bpnum, title, priority, status, mesg])
		if tsagent:
			status = "Pass"
			mesg = "Terminal Server Agent configured"
			ws.append([ip, bpnum, title, priority, status, mesg])
			for tsElement in entryElement.findall('ts-agent/entry'):
				profile = tsElement.attrib['name']
				host = tsElement.find('host')
				port = tsElement.find('port')
				status = "Informational"
				mesg = "Terminal Server Agent Profile %s configured for host %s:%s" % (profile, host.text, port.text)
				ws.append([ip, bpnum, title, priority, status, mesg])

	time.sleep(sleeptime)
	
def BP04018(ip, apikey):
	bpnum = "BP04018"
	title = "Disable WMI probing if not required."
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Management Profile
	xpath = "/config/devices/entry[@name='localhost.localdomain']/vsys"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	
	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall("./result/vsys/entry"):
		vsys = entryElement.attrib['name']
		for probeElement in entryElement.findall('user-id-collector/setting'):
			probeEnabled = probeElement.find('enable-probing')
			if not probeEnabled is None:
				status = "Informational"
				mesg = "WMI Probing is Enabled"
				ws.append([ip, bpnum, title, priority, status, mesg])

	time.sleep(sleeptime)
	
def BP04019(ip, apikey):
	bpnum = "BP04019"
	title = "Forbid User-ID on external and other non-trusted zones"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Interface Management Profile
	xpath = "/config/devices/entry[@name='localhost.localdomain']/network"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	responseElement = ET.fromstring(rrule.text)
	# Query for Interface Information
	xpath2 = "/config/devices/entry[@name='localhost.localdomain']/vsys"
	rulequery2 = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath2}
	rrule2 = requests.get('https://' + ip + '/api', params = rulequery2, verify=False)
	responseElement2 = ET.fromstring(rrule2.text)
	
	# Pull Management Profiles and Associated Interfaces
	for entryElement in responseElement.findall("./result/network"):
		for profileElement in entryElement.findall('profiles/interface-management-profile/entry'):
			profile = profileElement.attrib['name']
			uid = profileElement.find('userid-service')
			uidssl = profileElement.find('userid-syslog-listener-ssl')
			uidudp = profileElement.find('userid-syslog-listener-udp')
			uidprof = []
			if uid is None or uidssl is None or uidudp is None:
				continue
			elif uid.text == 'yes' or uidssl.text == 'yes' or uidudp.text == 'yes':
				uidprof.append(profile)
				for a in uidprof:
					status = "Informational"
					mesg = "Profile %s is configured for User-ID Services" % a
					ws.append([ip, bpnum, title, priority, status, mesg])

			for intfElement in entryElement.findall('interface/ethernet/entry'):
				intfprof = {}
				interface = intfElement.attrib['name']
				mgmtprofElement = intfElement.find('layer3/interface-management-profile')					
				if not mgmtprofElement is None and not interface in intfprof:
					intfprof[interface] = mgmtprofElement.text
					for interface, mgmtprofElement in intfprof.iteritems():
						if mgmtprofElement in uidprof:
							status = "Informational"
							mesg = "User-ID Management Profile %s is applied to interface %s" % (mgmtprofElement, interface)
							ws.append([ip, bpnum, title, priority, status, mesg])
	
	# Pull User Identification Enabled Zones
	for entryElement in responseElement2.findall("./result/vsys/entry"):
		vsys = entryElement.attrib['name']
		for zoneElement in entryElement.findall('zone/entry'):
			zone = zoneElement.attrib['name']
			uid = zoneElement.find('enable-user-identification')
			if uid is None:
				break
			elif uid.text == 'yes':
				status = "Informational"
				mesg = "User-ID Services are enabled for Zone %s" % zone
				ws.append([ip, bpnum, title, priority, status, mesg])

	time.sleep(sleeptime)

def BP04020(ip, apikey):
	bpnum = "BP04020"
	title = "Require the use of User-ID Include-Exclude Networks section, if User-ID is enabled."
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Management Profile
	xpath = "/config/devices/entry[@name='localhost.localdomain']/vsys"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	
	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall("./result/vsys/entry"):
		vsys = entryElement.attrib['name']
		uida = entryElement.find('user-id-agent')
		uidc = entryElement.find('user-id-collector')
		tsagent = entryElement.find('ts-agent')
		if uida is None and uidc is None and tsagent is None:
			continue
		if uida or uidc or tsagent:
			for useridElement in entryElement.findall('user-id-collector'):
				network = useridElement.find('include-exclude-network')
				if network is None:
					status = "Fail"
					mesg = "Network Include/Exclude not configured for system with User-ID enabled."
					ws.append([ip, bpnum, title, priority, status, mesg])
				else:
					for networkElement in useridElement.findall('include-exclude-network/entry'):
						entryname = networkElement.attrib['name']
						discovery = networkElement.find('discovery')
						network = networkElement.find('network-address')
						if re.match("(^10\.)|(^172\.1[6-9]\.)|(^172\.2[0-9]\.)|(^172\.3[0-1]\.)|(^192\.168\.)", network.text):
							status = "Pass"
							mesg = "User-ID Include/Exclude configured for %s network %s" % (discovery.text, network.text)
							ws.append([ip, bpnum, title, priority, status, mesg])
						else:
							status = "Informational"
							mesg = "User-ID Include/Exclude configured for %s network %s" % (discovery.text, network.text)
							ws.append([ip, bpnum, title, priority, status, mesg])

	time.sleep(sleeptime)
	
def BP04021(ip, apikey):
	bpnum = "BP04021"
	title = "Require default Log Forwarding Profile, this will be added automatically to all new Security Policies"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Local System Configuration
	xpath = '/config/shared'
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	responseElement = ET.fromstring(rrule.text)
	# Query for Panorama Configuration
	xpath2 = '/config/panorama'
	rulequery2 = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath2}
	rrule2 = requests.get('https://' + ip + '/api', params = rulequery2, verify=False)
	responseElement2 = ET.fromstring(rrule2.text)
	
	#Parse for Local System Configuration
	for entryElement in responseElement.findall('./result/shared/log-settings'):
		profiles = entryElement.find('profiles')
		defprof = []
		if profiles is None:
			status = "Informational"
			mesg = "No Local Log Forwarding Profile is Configured."
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif profiles:
			for profileElement in entryElement.findall('profiles/entry'):
				name = profileElement.attrib['name']
				defprof.append(name)
			if 'default' not in defprof:
				status = "Fail"
				mesg = "Default Log Forwarding Profile is Not Configured on the Local System."
				ws.append([ip, bpnum, title, priority, status, mesg])
			else:
				status = "Pass"
				mesg = "Default Log Forwarding Profile is Configured on the Local System."
				ws.append([ip, bpnum, title, priority, status, mesg])
				
	#Parse for Panorama Pushed Configuration
	for entryElement in responseElement2.findall('./result/panorama/vsys/entry'):
		vsys = entryElement.attrib['name']
		profiles = entryElement.find('log-settings/profiles')
		defprof = []
		if profiles is None:
			status = "Informational"
			mesg = "No Panorama Log Forwarding Profile is Configured."
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif profiles:
			for profileElement in entryElement.findall('log-settings/profiles/entry'):
				name = profileElement.attrib['name']
				defprof.append(name)
			if 'default' not in defprof:
				status = "Fail"
				mesg = "Default Log Forwarding Profile is Not pushed via Panorama Config."
				ws.append([ip, bpnum, title, priority, status, mesg])
			else:
				status = "Pass"
				mesg = "Default Log Forwarding Profile is pushed via Panorama Config."
				ws.append([ip, bpnum, title, priority, status, mesg])

	time.sleep(sleeptime)
	
def BP04022(ip, apikey):
	bpnum = "BP04022"
	title = "Increase WildFire file size upload limits"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Local System Configuration
	xpath = "/config/devices/entry[@name='localhost.localdomain']/deviceconfig/setting"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	responseElement = ET.fromstring(rrule.text)
		
	#Parse for Local System Configuration
	for entryElement in responseElement.findall('./result/setting'):
		wildfire = entryElement.find('wildfire')
		fivelimit = entryElement.find('wildfire/file-size-limit')
		if wildfire is None:
			status = "Fail"
			mesg = "Wildfire file upload sizes are set for the default."
			ws.append([ip, bpnum, title, priority, status, mesg])
		
		if fivelimit is None or fivelimit.text is None:
			pass
		elif fivelimit.text == '10':
			status = "Pass"
			mesg = "PANOS 5.0 Wildfire file size limit set to maximum of 10 MB"
			ws.append([ip, bpnum, title, priority, status, mesg])
		elif fivelimit.text != '10':
			status = "Fail"
			mesg = "PANOS 5.0 Wildfire file size limit set to %s" % fivelimit.text
			ws.append([ip, bpnum, title, priority, status, mesg])
			
		for sixonefileSize in entryElement.findall('wildfire/file-size-limit/entry'):
			filetype = sixonefileSize.attrib['name']
			sizelimit = sixonefileSize.find('size-limit')
			if filetype == 'flash' and sizelimit.text == '10':
				status = "Pass"
				mesg = "Flash file type maximum upload size is set to 10 MB"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif filetype == 'flash' and sizelimit.text != '10':
				status = "Fail"
				mesg = "Flash file type should be set to 10 MB, currently set to %s" % sizelimit.text
				ws.append([ip, bpnum, title, priority, status, mesg])
				
			if filetype == 'apk' and sizelimit.text == '50':
				status = "Pass"
				mesg = "APK file type maximum upload size is set to 50 MB"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif filetype == 'apk' and sizelimit.text != '50':
				status = "Fail"
				mesg = "APK file type should be set to 50 MB, currently set to %s" % sizelimit.text
				ws.append([ip, bpnum, title, priority, status, mesg])
				
			if filetype == 'pdf' and sizelimit.text == '1000':
				status = "Pass"
				mesg = "PDF file type maximum upload size is set to 1000 KB"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif filetype == 'pdf' and sizelimit.text != '1000':
				status = "Fail"
				mesg = "PDF file type should be set to 1000 KB, currently set to %s" % sizelimit.text
				ws.append([ip, bpnum, title, priority, status, mesg])
				
			if filetype == 'jar' and sizelimit.text == '10':
				status = "Pass"
				mesg = "JAR file type maximum upload size is set to 10 MB"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif filetype == 'jar' and sizelimit.text != '10':
				status = "Fail"
				mesg = "JAR file type should be set to 10 MB, currently set to %s" % sizelimit.text
				ws.append([ip, bpnum, title, priority, status, mesg])
				
			if filetype == 'pe' and sizelimit.text == '10':
				status = "Pass"
				mesg = "PE file type maximum upload size is set to 10 MB"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif filetype == 'pe' and sizelimit.text != '10':
				status = "Fail"
				mesg = "PE file type should be set to 10 MB, currently set to %s" % sizelimit.text
				ws.append([ip, bpnum, title, priority, status, mesg])	
				
			if filetype == 'ms-office' and sizelimit.text == '10000':
				status = "Pass"
				mesg = "MS Office file type maximum upload size is set to 10000 KB"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif filetype == 'ms-office' and sizelimit.text != '10000':
				status = "Fail"
				mesg = "MS Office file type should be set to 10000 KB, currently set to %s" % sizelimit.text
				ws.append([ip, bpnum, title, priority, status, mesg])	

		for sixohfileSize in entryElement.findall('wildfire/file-size-limit'):
			jar = sixohfileSize.find('jar')
			pe = sixohfileSize.find('pe')
			apk = sixohfileSize.find('apk')
			msoffice = sixohfileSize.find('ms-office')
			pdf = sixohfileSize.find('pdf')
			if apk is None:
				pass
			elif apk.text == '50':
				status = "Pass"
				mesg = "APK file type maximum upload size is set to 50 MB"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif apk.text != '50':
				status = "Fail"
				mesg = "APK file type should be set to 50 MB, currently set to %s" % apk.text
				ws.append([ip, bpnum, title, priority, status, mesg])
				
			if pdf is None:
				pass
			elif pdf.text == '1000':
				status = "Pass"
				mesg = "PDF file type maximum upload size is set to 1000 KB"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif pdf.text != '1000':
				status = "Fail"
				mesg = "PDF file type should be set to 1000 KB, currently set to %s" % pdf.text
				ws.append([ip, bpnum, title, priority, status, mesg])
				
			if jar is None:
				pass
			elif jar.text == '10':
				status = "Pass"
				mesg = "JAR file type maximum upload size is set to 10 MB"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif jar.text != '10':
				status = "Fail"
				mesg = "JAR file type should be set to 10 MB, currently set to %s" % jar.text
				ws.append([ip, bpnum, title, priority, status, mesg])
				
			if pe is None:
				pass
			elif pe.text == '10':
				status = "Pass"
				mesg = "PE file type maximum upload size is set to 10 MB"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif pe.text != '10':
				status = "Fail"
				mesg = "PE file type should be set to 10 MB, currently set to %s" % pe.text
				ws.append([ip, bpnum, title, priority, status, mesg])	
				
			if msoffice is None:
				pass
			elif msoffice.text == '10000':
				status = "Pass"
				mesg = "MS Office file type maximum upload size is set to 10000 KB"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif msoffice.text != '10000':
				status = "Fail"
				mesg = "MS Office file type should be set to 10000 KB, currently set to %s" % msoffice.text
				ws.append([ip, bpnum, title, priority, status, mesg])

	time.sleep(sleeptime)
	
def BP04023(ip, apikey):
	bpnum = "BP04023"
	title = "Require WildFire File Blocking profiles to include any application, any file type, and action set to forward"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Local System Configuration
	xpath = "/config/devices/entry[@name='localhost.localdomain']/vsys"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	responseElement = ET.fromstring(rrule.text)
	# Query for Panorama Configuration
	xpath2 = "/config/panorama"
	rulequery2 = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath2}
	rrule2 = requests.get('https://' + ip + '/api', params = rulequery2, verify=False)
	responseElement2 = ET.fromstring(rrule2.text)
	
	#Parse for Local System Configuration
	for entryElement in responseElement.findall('./result/vsys/entry'):
		vsys = entryElement.attrib['name']
		for fbprofElement in entryElement.findall('profiles/file-blocking/entry'):
			fbprofile = fbprofElement.attrib['name']
			for fbruleElement in fbprofElement.findall('rules/entry'):
				fbrule = fbruleElement.attrib['name']
				fbaction = fbruleElement.find('action')
				for appElement in fbruleElement.findall('application'):
					app = appElement.find('member')
				for typeElement in fbruleElement.findall('file-type'):
					filetype = typeElement.find('member')
					if fbaction.text == 'forward' and app.text == 'any' and filetype.text == 'any':
						status = "Pass"
						mesg = "File Blocking Profile '%s',rule '%s' is configured with Action forward and App and Filetype of Any" % (fbprofile, fbrule)
						ws.append([ip, bpnum, title, priority, status, mesg])
					elif fbaction.text == 'forward' and (app.text != 'any' or filetype.text != 'any'):
						status = "Fail"
						mesg = "File Blocking Profile '%s', rule '%s' is configured with Action forward and custom App or Filetype" % (fbprofile, fbrule)
						ws.append([ip, bpnum, title, priority, status, mesg])
						
	for entryElement in responseElement.findall('./result/vsys/entry'):
		vsys = entryElement.attrib['name']
		for fbprofElement in entryElement.findall('profiles/wildfire-analysis/entry'):
			fbprofile = fbprofElement.attrib['name']
			for fbruleElement in fbprofElement.findall('rules/entry'):
				fbrule = fbruleElement.attrib['name']
				for appElement in fbruleElement.findall('application'):
					app = appElement.find('member')
				for typeElement in fbruleElement.findall('file-type'):
					filetype = typeElement.find('member')
					if app.text == 'any' and filetype.text == 'any':
						status = "Pass"
						mesg = "Wildfire Analysis Profile '%s',rule '%s' is configured with App and Filetype of Any" % (fbprofile, fbrule)
						ws.append([ip, bpnum, title, priority, status, mesg])
					elif app.text != 'any' or filetype.text != 'any':
						status = "Fail"
						mesg = "Wildfire Analysis Profile '%s', rule '%s' is configured with custom App or Filetype" % (fbprofile, fbrule)
						ws.append([ip, bpnum, title, priority, status, mesg])	
					
					
	#Parse for Panorama Defined Configuration
	for entryElement in responseElement2.findall('./result/panorama/vsys/entry'):
		vsys = entryElement.attrib['name']
		for fbprofElement in entryElement.findall('profiles/file-blocking/entry'):
			fbprofile = fbprofElement.attrib['name']
			for fbruleElement in fbprofElement.findall('rules/entry'):
				fbrule = fbruleElement.attrib['name']
				fbaction = fbruleElement.find('action')
				for appElement in fbruleElement.findall('application'):
					app = appElement.find('member')
				for typeElement in fbruleElement.findall('file-type'):
					filetype = typeElement.find('member')
					if fbaction.text == 'forward' and app.text == 'any' and filetype.text == 'any':
						status = "Pass"
						mesg = "File Blocking Profile '%s',rule '%s' is configured with Action forward and App and Filetype of Any" % (fbprofile, fbrule)
						ws.append([ip, bpnum, title, priority, status, mesg])
					elif fbaction.text == 'forward' and (app.text != 'any' or filetype.text != 'any'):
						status = "Fail"
						mesg = "File Blocking Profile '%s', rule '%s' is configured with Action forward and custom App or Filetype" % (fbprofile, fbrule)
						ws.append([ip, bpnum, title, priority, status, mesg])
						
	for entryElement in responseElement2.findall('./result/panorama/vsys/entry'):
		vsys = entryElement.attrib['name']
		for fbprofElement in entryElement.findall('profiles/wildfire-analysis/entry'):
			fbprofile = fbprofElement.attrib['name']
			for fbruleElement in fbprofElement.findall('rules/entry'):
				fbrule = fbruleElement.attrib['name']
				for appElement in fbruleElement.findall('application'):
					app = appElement.find('member')
				for typeElement in fbruleElement.findall('file-type'):
					filetype = typeElement.find('member')
					if app.text == 'any' and filetype.text == 'any':
						status = "Pass"
						mesg = "Wildfire Analysis Profile '%s',rule '%s' is configured with App and Filetype of Any" % (fbprofile, fbrule)
						ws.append([ip, bpnum, title, priority, status, mesg])
					elif app.text != 'any' or filetype.text != 'any':
						status = "Fail"
						mesg = "Wildfire Analysis Profile '%s', rule '%s' is configured with custom App or Filetype" % (fbprofile, fbrule)
						ws.append([ip, bpnum, title, priority, status, mesg])

	time.sleep(sleeptime)

def BP04025(ip, apikey):
	bpnum = "BP04025"
	title = "Require forwarding of decrypted content"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Local System Configuration
	xpath = "/config/devices/entry[@name='localhost.localdomain']/vsys"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	responseElement = ET.fromstring(rrule.text)
	
	#Parse for Local System Configuration
	for entryElement in responseElement.findall('./result/vsys/entry'):
		vsys = entryElement.attrib['name']
		check = entryElement.find('setting')
		if check is None:
			status = "Fail"
			mesg = "Forwarding of SSL Decrypted Content is Disabled"
			ws.append([ip, bpnum, title, priority, status, mesg])
		for settingElement in entryElement.findall('setting'):
			fwddecrypt = settingElement.find('ssl-decrypt/allow-forward-decrypted-content')
			if fwddecrypt is None or fwddecrypt.text == 'no':
				status = "Fail"
				mesg = "Forwarding of SSL Decrypted Content is Disabled"
				ws.append([ip, bpnum, title, priority, status, mesg])
			elif fwddecrypt.text == 'yes':
				status = "Pass"
				mesg = "Forwarding of SSL Decrypted Content is Enabled."
				ws.append([ip, bpnum, title, priority, status, mesg])

	time.sleep(sleeptime)
	
#-------Rule Definitions------
#----Rule 08000 - 9999: Panorama Platform Specific Rules
def BP08000(ip, apikey):
	bpnum = "BP08000"
	title = "Panorama Platform eTAC Recommended Versions of Code"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Management Profile
	xpath = "<show><system><info></info></system></show>"
	rulequery = {'type': 'op', 'action': 'get', 'key': apikey, 'cmd': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	
	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall("./result/system"):
		version = entryElement.find('sw-version')
		if version.text == "6.0.8" or version.text == "6.0.9" or version.text == "6.0.10" or version.text == "6.1.4" or version.text == "6.1.5":
			status = "Pass"
			mesg = "Current Version: %s - Firewall is running an eTAC Recommended Version of Code" % version.text
			ws.append([ip, bpnum, title, priority, status, mesg])
		else:
			status = "Fail"
			mesg = "Current Version: %s - Firewall is not running an eTAC Recommended Version of Code. For more information refer to https://intranet.paloaltonetworks.com/docs/DOC-4857" % version.text
			ws.append([ip, bpnum, title, priority, status, mesg])
			
	time.sleep(sleeptime)

def BP08001(ip, apikey):	
	bpnum = "BP08001"
	title = "Panorama Platform Should Use RADIUS for User Authentication"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Configured Authentication Profiles
	xpath = "/config/devices/entry[@name='localhost.localdomain']/deviceconfig/system"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	# Query for Authentication Profile Protocol
	xpath2 = "/config/panorama"
	rulequery2 = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath2}
	rrule2 = requests.get('https://' + ip + '/api', params = rulequery2, verify=False)
		
	responseElement = ET.fromstring(rrule.text) 
	for entryElement in responseElement.findall("./result/system"):
		sysauth = entryElement.find('authentication-profile')
		sysauthprof = []
		if sysauth is None:
			status = "Fail"
			mesg = "Authentication Profile is not Configured, Using local authentication."
			ws.append([ip, bpnum, title, priority, status, mesg])
		else:
			sysauthprof.append(sysauth.text)
			status = "Informational"
			mesg = "Authentication Profile %s is configured" % sysauth.text
			ws.append([ip, bpnum, title, priority, status, mesg])
		
		responseElement2 = ET.fromstring(rrule2.text)
		for authprofElement in responseElement2.findall("./result/panorama/authentication-profile"):
			for entryElement in authprofElement.findall('entry'):
				entryName = entryElement.attrib['name']
				if entryName in sysauthprof:
					for methodElement in entryElement.findall('method'):
						for methodTypeElement in methodElement:
							methodType = methodTypeElement.tag.upper()
							status = "Pass"
							mesg = "Panorama Authentication Profile %s is using %s" % (entryName , methodType)
							ws.append([ip, bpnum, title, priority, status, mesg])

	time.sleep(sleeptime)
	


def BP08002(ip, apikey):
	bpnum = "BP08002"
	title = "Panorama Platform Define Syslog Servers"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Shared Config
	xpath = "/config/panorama"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)

	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall('./result/panorama'):
		syslog = entryElement.find('log-settings/syslog')
		if syslog is None:
			status = "Fail"
			mesg = "Syslog not configured on system"
			ws.append([ip, bpnum, title, priority, status, mesg])
		else:
			for logsetElement in entryElement.findall('log-settings/syslog/entry'):
				profile = logsetElement.attrib['name']
				for srvElement in logsetElement.findall('server/entry'):
					server = srvElement.attrib['name']
					facility = srvElement.find('facility')
					ipaddr = srvElement.find('server')
					if facility.text == 'LOG_USER':
						status = "Pass"
						mesg = "SYSLOG Server %s is configured for IP %s using facility %s" % (server, ipaddr.text, facility.text)
						ws.append([ip, bpnum, title, priority, status, mesg])
					elif facility.text != 'LOG_USER':
						status = "Fail"
						mesg = "SYSLOG Server %s is configured for IP %s using facility %s" % (server, ipaddr.text, facility.text)
						ws.append([ip, bpnum, title, priority, status, mesg])
						
	time.sleep(sleeptime)

def BP08003(ip, apikey):
	bpnum = "BP08003"
	title = "Panorama Platform Define SNMP TRAP Servers"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Shared Config
	xpath = "/config/panorama"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)

	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall('./result/panorama'):
		snmpTrap = entryElement.find('log-settings/snmptrap')
		if snmpTrap is None:
			status = "Fail"
			mesg = "SNMP Trap Server not configured on system"
			ws.append([ip, bpnum, title, priority, status, mesg])
		else:
			for snmpElement in entryElement.findall('log-settings/snmptrap/entry'):
				profile = snmpElement.attrib['name']
				for v2srvElement in snmpElement.findall('version/v2c/server/entry'):
					server = v2srvElement.attrib['name']
					ipaddr = v2srvElement.find('manager')
					status = "Pass"
					mesg = "SNMP v2 Trap Server %s is configured for IP %s" % (server, ipaddr.text)
					ws.append([ip, bpnum, title, priority, status, mesg])
				for v3srvElement in snmpElement.findall('version/v3/server/entry'):
					server = v3srvElement.attrib['name']
					ipaddr = v3srvElement.find('manager')
					status = "Pass"
					mesg = "SNMP v3 Trap Server %s is configured for IP %s" % (server, ipaddr.text)
					ws.append([ip, bpnum, title, priority, status, mesg])
	
	time.sleep(sleeptime)

def BP08004(ip, apikey):
	bpnum = "BP08004"
	title = "Panorama Platform Configure target e-mail account to receive critical and high severity threat, system, and Wildfire log events."
	priority = "Informational"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Panorama Config
	xpath = "/config/panorama"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)

	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall('./result/panorama'):
		email = entryElement.find('log-settings/email/')
		if email is None:
			status = "Informational"
			mesg = "Email Profile is not configured on system"
			ws.append([ip, bpnum, title, priority, status, mesg])
		else:
			for emailsetElement in entryElement.findall('log-settings/email/entry'):
				profile = emailsetElement.attrib['name']
				for srvElement in emailsetElement.findall('server/entry'):
					server = srvElement.attrib['name']
					gateway = srvElement.find('gateway')
					to = srvElement.find('to')
					status = "Informational"
					mesg = "Email Profile %s is configured for gateway %s using email %s" % (profile, gateway.text, to.text)
					ws.append([ip, bpnum, title, priority, status, mesg])
					
	time.sleep(sleeptime)
	
def BP08005(ip, apikey):
	bpnum = "BP08005"
	title = "Configure Panorama Platform System Logs to Forward to Syslog, SNMP, or eMail."
	priority = "Medium"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Panorama Config
	xpath = "/config/panorama"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)

	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall('./result/panorama'):
		system = entryElement.find('log-settings/system/')
		informational = entryElement.find('log-settings/system/informational')
		low = entryElement.find('log-settings/system/low')
		medium = entryElement.find('log-settings/system/medium')
		high = entryElement.find('log-settings/system/high')
		critical = entryElement.find('log-settings/system/critical')
		if system is None:
			status = "Fail"
			mesg = "System Logging is not configured."
			ws.append([ip, bpnum, title, priority, status, mesg])
			break
		
		if not informational is None:
			for infoElement in entryElement.findall('log-settings/system/informational'):
				loggrp = {}
				snmp = infoElement.find('send-snmptrap/using-snmptrap-setting')
				email = infoElement.find('send-email/using-email-setting')
				syslog = infoElement.find('send-syslog/using-syslog-setting')
				if not snmp is None:
					loggrp['snmp'] = snmp.text
				if not email is None:
					loggrp['email'] = email.text
				if not syslog is None:
					loggrp['syslog'] = syslog.text
				for type, profile in loggrp.iteritems():
					status = "Pass"
					mesg = "System Informational Logging configured for " + type + ": " + profile
					ws.append([ip, bpnum, title, priority, status, mesg])
		else:
			status = "Informational"
			mesg = "Informational Logging Profiles not Configured."
			ws.append([ip, bpnum, title, priority, status, mesg])

		
		if not low is None:
			for infoElement in entryElement.findall('log-settings/system/low'):
				loggrp = {}
				snmp = infoElement.find('send-snmptrap/using-snmptrap-setting')
				email = infoElement.find('send-email/using-email-setting')
				syslog = infoElement.find('send-syslog/using-syslog-setting')
				if not snmp is None:
					loggrp['snmp'] = snmp.text
				if not email is None:
					loggrp['email'] = email.text
				if not syslog is None:
					loggrp['syslog'] = syslog.text
				for type, profile in loggrp.iteritems():
					status = "Pass"
					mesg = "System Low Logging configured for " + type + ": " + profile
					ws.append([ip, bpnum, title, priority, status, mesg])
		else:
			status = "Informational"
			mesg = "Low Logging Profiles not Configured."
			ws.append([ip, bpnum, title, priority, status, mesg])
			
		if not medium is None:
			for infoElement in entryElement.findall('log-settings/system/medium'):
				loggrp = {}
				snmp = infoElement.find('send-snmptrap/using-snmptrap-setting')
				email = infoElement.find('send-email/using-email-setting')
				syslog = infoElement.find('send-syslog/using-syslog-setting')
				if not snmp is None:
					loggrp['snmp'] = snmp.text
				if not email is None:
					loggrp['email'] = email.text
				if not syslog is None:
					loggrp['syslog'] = syslog.text
				for type, profile in loggrp.iteritems():
					status = "Pass"
					mesg = "System Medium Logging configured for " + type + ": " + profile
					ws.append([ip, bpnum, title, priority, status, mesg])
		else:
			status = "Fail"
			mesg = "Medium Logging Profiles not Configured."
			ws.append([ip, bpnum, title, priority, status, mesg])
		
		if not high is None:
			for infoElement in entryElement.findall('log-settings/system/high'):
				loggrp = {}
				snmp = infoElement.find('send-snmptrap/using-snmptrap-setting')
				email = infoElement.find('send-email/using-email-setting')
				syslog = infoElement.find('send-syslog/using-syslog-setting')
				if not snmp is None:
					loggrp['snmp'] = snmp.text
				if not email is None:
					loggrp['email'] = email.text
				if not syslog is None:
					loggrp['syslog'] = syslog.text
				for type, profile in loggrp.iteritems():
					status = "Pass"
					mesg = "System High Logging configured for " + type + ": " + profile
					ws.append([ip, bpnum, title, priority, status, mesg])
		else:
			status = "Fail"
			mesg = "High Logging Profiles not Configured."
			ws.append([ip, bpnum, title, priority, status, mesg])
				
		if not critical is None:
			for infoElement in entryElement.findall('log-settings/system/critical'):
				loggrp = {}
				snmp = infoElement.find('send-snmptrap/using-snmptrap-setting')
				email = infoElement.find('send-email/using-email-setting')
				syslog = infoElement.find('send-syslog/using-syslog-setting')
				if not snmp is None:
					loggrp['snmp'] = snmp.text
				if not email is None:
					loggrp['email'] = email.text
				if not syslog is None:
					loggrp['syslog'] = syslog.text
				for type, profile in loggrp.iteritems():
					status = "Pass"
					mesg = "System Critical Logging configured for " + type + ": " + profile
					ws.append([ip, bpnum, title, priority, status, mesg])
		else:
			status = "Fail"
			mesg = "Critical Logging Profiles not Configured."
			ws.append([ip, bpnum, title, priority, status, mesg])		

	time.sleep(sleeptime)

def BP08006(ip, apikey):	
	bpnum = "BP08006"
	title = "Configure Panorama Platform Config Logs to Forward to Syslog, SNMP, or eMail."
	priority = "Medium"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Panorama Config
	xpath = "/config/panorama"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)

	responseElement = ET.fromstring(rrule.text)
	for entryElement in responseElement.findall('./result/panorama'):
		config = entryElement.find('log-settings/config/')
		if config is None:
			status = "Fail"
			mesg = "Configuration Change Logging is not configured."
			ws.append([ip, bpnum, title, priority, status, mesg])		
			break
		
		for infoElement in entryElement.findall('log-settings/config/any'):
			loggrp = {}
			snmp = infoElement.find('send-snmptrap/using-snmptrap-setting')
			email = infoElement.find('send-email/using-email-setting')
			syslog = infoElement.find('send-syslog/using-syslog-setting')
			panorama = infoElement.find('send-to-panorama')
			if not snmp is None:
				loggrp['snmp'] = snmp.text
			if not email is None:
				loggrp['email'] = email.text
			if not syslog is None:
				loggrp['syslog'] = syslog.text
			if not panorama is None:
				loggrp['Panorama'] = panorama.text
			for type, profile in loggrp.iteritems():
				status = "Pass"
				mesg = "Configuration Change Logging configured for " + type + ": " + profile
				ws.append([ip, bpnum, title, priority, status, mesg])		
	
	time.sleep(sleeptime)


#-------Rule Definitions------
#----Rule 10000 - 13999: Panorama Managed Device Specific Rules
def BP10000(ip, apikey):
	bpnum = "BP10000"
	title = "Panorama Device Groups Use Descriptive Rule Names"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Device Group Rules
	xpath = "/config/devices/entry[@name='localhost.localdomain']/device-group"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	
	# Query for Shared Rules
	xpath2 = "/config/shared"
	rulequery2 = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath2}
	rrule2 = requests.get('https://' + ip + '/api', params = rulequery2, verify=False)

	responseElement = ET.fromstring(rrule.content)
	for entryElement in responseElement.findall("./result/device-group/entry"):
		devgrp = entryElement.attrib['name']
		for secruleElement in entryElement.findall('pre-rulebase/security/rules/entry'):
			rule = secruleElement.attrib['name']
			if re.match("pre-rule\d{1,2}", rule):
				status = "Fail"
				mesg = "Pre-Rule %s in %s is using a default rule name" % (rule, devgrp)
				ws.append([ip, bpnum, title, priority, status, mesg])
		for secruleElement in entryElement.findall('post-rulebase/security/rules/entry'):
			rule = secruleElement.attrib['name']
			if re.match("post-rule\d", rule):
				status = "Fail"
				mesg = "Post-Rule %s in %s is using a default rule name" % (rule, devgrp)
				ws.append([ip, bpnum, title, priority, status, mesg])
	
	responseElement2 = ET.fromstring(rrule2.content)
	for entryElement in responseElement2.findall("./result/shared/pre-rulebase/security/rules/entry"):
		rule = entryElement.attrib['name']
		if re.match("shared-pre-rule\d", rule):
			status = "Fail"
			mesg = "Shared Pre-Rule %s is using a default rule name" % rule
			ws.append([ip, bpnum, title, priority, status, mesg])
	for entryElement in responseElement2.findall("./result/shared/post-rulebase/security/rules/entry"):
		rule = entryElement.attrib['name']
		if re.match("shared-post-rule\d", rule):
			status = "Fail"
			mesg = "Shared Post-Rule %s is using a default rule name" % rule
			ws.append([ip, bpnum, title, priority, status, mesg])
			
	time.sleep(sleeptime)
	
def BP10001(ip, apikey):
	bpnum = "BP10001"
	title = "Panorama Device Groups Alert on Invalid Subnet Masks"
	priority = "Low"
	print "Running Rule %s - %s" % (bpnum, title)
	# Query for Device Group Objects
	xpath = "/config/devices/entry[@name='localhost.localdomain']/device-group"
	rulequery = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath}
	rrule = requests.get('https://' + ip + '/api', params = rulequery, verify=False)
	# Query for Shared Objects
	xpath2 = "/config/shared"
	rulequery2 = {'type': 'config', 'action': 'get', 'key': apikey, 'xpath': xpath2}
	rrule2 = requests.get('https://' + ip + '/api', params = rulequery2, verify=False)

	responseElement = ET.fromstring(rrule.content)
	for entryElement in responseElement.findall('./result/device-group/entry'):
		devgrp = entryElement.attrib['name']
		## Check Device Group Pre-Rulebase for Invalid Mask in rule
		for secruleElement in entryElement.findall('pre-rulebase/security/rules/entry'):
			rule = secruleElement.attrib['name']
			for srcaddrElement in secruleElement.findall('source'):
				srcaddr = srcaddrElement.find('member')
				if srcaddr is None:
					continue
				elif re.match("^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\/([1-7]{1})$", srcaddr.text):
					status = "Fail"
					mesg = "Invalid Subnet Mask %s found in Device Group '%s' Rule '%s'" % (srcaddr.text, devgrp, rule)
					ws.append([ip, bpnum, title, priority, status, mesg])
			for dstaddrElement in secruleElement.findall('destination'):
				dstaddr = dstaddrElement.find('member')
				if dstaddr is None:
					continue
				elif re.match("^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\/([1-7]{1})$", dstaddr.text):
					status = "Fail"
					mesg = "Invalid Subnet Mask %s found in Device Group '%s' Rule '%s'" % (dstaddr.text, devgrp, rule)
					ws.append([ip, bpnum, title, priority, status, mesg])
		## Check Device Group Post-Rulebase for Invalid Mask in rule
		for secruleElement in entryElement.findall('post-rulebase/security/rules/entry'):
			rule = secruleElement.attrib['name']
			for srcaddrElement in secruleElement.findall('source'):
				for srcaddr in srcaddrElement.findall('member'):
					if srcaddr is None:
						continue
					elif re.match("^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\/([1-7]{1})$", srcaddr.text):
						status = "Fail"
						mesg = "Invalid Subnet Mask %s found in Device Group '%s' Rule '%s'" % (srcaddr.text, devgrp, rule)
						ws.append([ip, bpnum, title, priority, status, mesg])
			for dstaddrElement in secruleElement.findall('destination'):
				for dstaddr in dstaddrElement.findall('member'):
					if dstaddr is None:
						continue
					elif re.match("^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\/([1-7]{1})$", dstaddr.text):
						status = "Fail"
						mesg = "Invalid Subnet Mask %s found in Device Group '%s' Rule '%s'" % (dstaddr.text, devgrp, rule)
						ws.append([ip, bpnum, title, priority, status, mesg])

		## Check Device Group Address Objects for Invalid Mask						
		for devaddrElement in entryElement.findall('address/entry'):
			address = devaddrElement.attrib['name']
			ipmask = devaddrElement.find('ip-netmask')
			if ipmask is None:
				continue
			elif re.match("^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\/([1-7]{1})$", ipmask.text):
				status = "Fail"
				mesg = "Invalid Subnet Mask %s found in Device Group '%s' Address Object '%s'" % (ipmask.text, devgrp, address)
				ws.append([ip, bpnum, title, priority, status, mesg])
			
	
	responseElement2 = ET.fromstring(rrule2.content)
	for entryElement in responseElement2.findall('./result/shared'):
		## Check Shared Pre-Rulebase for Invalid Mask in rule
		for secruleElement in entryElement.findall('pre-rulebase/security/rules/entry'):
			rule = secruleElement.attrib['name']
			for srcaddrElement in secruleElement.findall('source'):
				srcaddr = srcaddrElement.find('member')
				if srcaddr is None:
					continue
				elif re.match("^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\/([1-7]{1})$", srcaddr.text):
					status = "Fail"
					mesg = "Invalid Subnet Mask %s found in Shared Rule '%s'" % (srcaddr.text, rule)
					ws.append([ip, bpnum, title, priority, status, mesg])
			for dstaddrElement in secruleElement.findall('destination'):
				dstaddr = dstaddrElement.find('member')
				if dstaddr is None:
					continue
				elif re.match("^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\/([1-7]{1})$", dstaddr.text):
					status = "Fail"
					mesg = "Invalid Subnet Mask %s found in Shared Rule '%s'" % (dstaddr.text, rule)
					ws.append([ip, bpnum, title, priority, status, mesg])
		## Check Shared Post-Rulebase for Invalid Mask in rule
		for secruleElement in entryElement.findall('post-rulebase/security/rules/entry'):
			rule = secruleElement.attrib['name']
			for srcaddrElement in secruleElement.findall('source'):
				for srcaddr in srcaddrElement.findall('member'):
					if srcaddr is None:
						continue
					elif re.match("^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\/([1-7]{1})$", srcaddr.text):
						status = "Fail"
						mesg = "Invalid Subnet Mask %s found in Shared Rule '%s'" % (srcaddr.text, rule)
						ws.append([ip, bpnum, title, priority, status, mesg])
			for dstaddrElement in secruleElement.findall('destination'):
				for dstaddr in dstaddrElement.findall('member'):
					if dstaddr is None:
						continue
					elif re.match("^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\/([1-7]{1})$", dstaddr.text):
						status = "Fail"
						mesg = "Invalid Subnet Mask %s found in Shared Rule '%s'" % (dstaddr.text, rule)
						ws.append([ip, bpnum, title, priority, status, mesg])

		## Check Shared Address Objects for Invalid Mask							
		for srdaddrElement in entryElement.findall('address/entry'):
			address = srdaddrElement.attrib['name']
			ipmask = srdaddrElement.find('ip-netmask')
			if ipmask is None:
				continue
			elif re.match("^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\/([1-7]{1})$", ipmask.text):
				status = "Fail"
				mesg = "Invalid Subnet Mask %s found in Shared Address Object '%s'" % (ipmask.text, address)
				ws.append([ip, bpnum, title, priority, status, mesg])
			
	time.sleep(sleeptime)
					
	
#-------Device Group Definitions------
#----Defines which rules are tied to which device class.
	
def BPPanorama(ip, apikey):
	BP01000(ip, apikey)  #  Configure Hostname on System
	BP01001(ip, apikey)  #  Configure Domain Name on System
	BP01002(ip, apikey)  #  Replace default administrative super-user account by creating a new superuser account and then deleting the admin account.
	BP01003(ip, apikey)  #  Configure administrative access lockouts and timeouts in the authentication settings portion of the device setup tab. Configure administrative timeout to 10 minutes; apply these settings to administrator accounts.
	BP01005(ip, apikey)  #  Configure the firewall to verify the identity of the Palo Alto update server on the device setup page.
	BP01006(ip, apikey)  #  Configure login banner.
	BP01007(ip, apikey)  #  Configure Geo-Location Longitude and Latitude
	BP01008(ip, apikey)  #  Configure the firewall to use redundant DNS Servers
	BP01009(ip, apikey)  #  Configure the firewall to use redundant NTP Servers
	BP01010(ip, apikey)  #  Limit management interface traffic to ping and secure protocols only
	BP01011(ip, apikey)  #  Limit permitted IP Addresses to those necessary for device management
	BP01012(ip, apikey)
	BP01015(ip, apikey)  #  Configure minimum password complexity profile 
	BP01016(ip, apikey)  #  Ensure that Firewall Dynamic Updates are Processing
	BP01017(ip, apikey)  #  Ensure that Dynamic Update Times are Configured Properly
	BP01018(ip, apikey)
	BP08000(ip, apikey)  #  Panorama Platform eTAC Recommended Versions of Code
	BP08001(ip, apikey)  #  Panorama Platform Should Use RADIUS for User Authentication
	BP08002(ip, apikey)  #  Panorama Platform Define Syslog Servers
	BP08003(ip, apikey)  #  Panorama Platform Define SNMP TRAP Servers
	BP08004(ip, apikey)  #  Panorama Platform Configure target e-mail account to receive critical and high severity threat, system, and Wildfire log events.
	BP08005(ip, apikey)  #  Configure Panorama Platform System Logs to Forward to Syslog, SNMP, or eMail.
	BP08006(ip, apikey)  #  Configure Panorama Platform Config Logs to Forward to Syslog, SNMP, or eMail.
	BP10000(ip, apikey)  
	BP10001(ip, apikey)  

def BPUmgPan(ip, apikey):
	BP01000(ip, apikey)
	BP01001(ip, apikey)
	BP01002(ip, apikey)
	BP01003(ip, apikey)
	BP01004(ip, apikey)
	BP01005(ip, apikey)
	BP01006(ip, apikey)
	BP01007(ip, apikey)
	BP01008(ip, apikey)
	BP01009(ip, apikey)
	BP01010(ip, apikey)
	BP01011(ip, apikey)
	BP01012(ip, apikey)
	BP01015(ip, apikey)
	BP01016(ip, apikey)
	BP01017(ip, apikey)
	BP01018(ip, apikey)
	BP04000(ip, apikey)  #  Firewall Should Use Descriptive Rule Names
	BP04001(ip, apikey)  #  Firewall eTAC Recommended Versions of Code
	BP04002(ip, apikey)  #  Firewall Should Use RADIUS for User Authentication
	BP04003(ip, apikey)  #  Firewall Alert on Invalid Subnet Masks
	BP04004(ip, apikey)  #  Define Syslog Servers on Firewall
	BP04005(ip, apikey)  #  Define SNMP TRAP Servers on Firewall
	BP04006(ip, apikey)  #  Define NetFlow Collectors on Firewall
	BP04007(ip, apikey)  #  Validate That a NetFlow Profile is Assigned to an Interface
	BP04008(ip, apikey)  #  Configure target e-mail account to receive critical and high severity threat, system, and Wildfire log events.
	BP04009(ip, apikey)  #  Configure Firewall System Logs to Forward to Panorama, Syslog, SNMP, or eMail.
	BP04010(ip, apikey)  #  Configure Firewall Config Logs to Forward to Panorama, Syslog, SNMP, or eMail.
	BP04011(ip, apikey)
	BP04012(ip, apikey)
	BP04013(ip, apikey)
	BP04014(ip, apikey)
	BP04015(ip, apikey)
	BP04016(ip, apikey)
	BP04017(ip, apikey)
	BP04018(ip, apikey)
	BP04019(ip, apikey)
	BP04020(ip, apikey)
	BP04021(ip, apikey)
	BP04022(ip, apikey)
	BP04023(ip, apikey)
	BP04025(ip, apikey)

def BPMGPan(ip, apikey):
	BP01000(ip, apikey)
	BP01001(ip, apikey)
	BP01002(ip, apikey)
	BP01003(ip, apikey)
	BP01004(ip, apikey)
	BP01005(ip, apikey)
	BP01006(ip, apikey)
	BP01007(ip, apikey)
	BP01008(ip, apikey)
	BP01009(ip, apikey)
	BP01010(ip, apikey)
	BP01011(ip, apikey)
	BP01012(ip, apikey)
	BP01015(ip, apikey)
	BP01016(ip, apikey)
	BP01017(ip, apikey)
	BP01018(ip, apikey)
	BP04000(ip, apikey)
	BP04001(ip, apikey)
	BP04002(ip, apikey)
	BP04003(ip, apikey)
	BP04004(ip, apikey)
	BP04005(ip, apikey)
	BP04006(ip, apikey)
	BP04007(ip, apikey)
	BP04008(ip, apikey)
	BP04009(ip, apikey)
	BP04010(ip, apikey)
	BP04011(ip, apikey)
	BP04012(ip, apikey)
	BP04013(ip, apikey)
	BP04014(ip, apikey)
	BP04015(ip, apikey)
	BP04016(ip, apikey)
	BP04017(ip, apikey)
	BP04018(ip, apikey)
	BP04019(ip, apikey)
	BP04020(ip, apikey)
	BP04021(ip, apikey)
	BP04022(ip, apikey)
	BP04023(ip, apikey)
	BP04025(ip, apikey)
	
def	BPTool():
	with open(fwinfo, 'r+') as csvfile:
		reader = csv.DictReader(csvfile)
		for row in reader:
			ip = row["ip_address"]
			keygen = {'type': 'keygen', 'user': row["username"], 'password': row["password"]}
			rkey = requests.post('https://' + row["ip_address"] + '/api', params=keygen, verify=False)
			rresp = ET.fromstring(rkey.content)
			apikeysearch = rresp.find('result/key')
			apikey = apikeysearch.text
			print "Generating API Key for %s" % ip
					
			#lookup device type and invoke aggregating function
			devicetype = row["device_type"]
			if devicetype == "Panorama":
				BPPanorama(ip, apikey)
			elif devicetype == "Managed-PAN":
				BPMGPan(ip, apikey)
			elif devicetype == "Unmanaged-PAN":
				BPUmgPan(ip, apikey)
			else:
				print "Device Type Not Set. Should be Panorama, Managed-PAN, or Unmanaged-PAN"
				

	
	
	
	
print ""
print "##############################################################"
print "#### Palo Alto Best Practices Analysis Tool               ####"
print "#### 				                         ####"
print "#### Version: 0.7 ALPHA (Evergreen)                       ####"
print "####                                                      ####"
print "#### Written By: Jessica Ferguson                         ####"
print "#### c0verfire@hush.com		                         ####"
print "####			                                 ####"
print "##############################################################"
print ""
print "This software should be considered alpha code and should not be used in a production environment"
print "Here there be Dragons!!! Proceed at your Own Risk"
print ""

proceed = raw_input("Enter (y)es or (n)o: ") 
if proceed == "yes" or proceed == "y": 
	cust = raw_input("Please enter the customer name: ")
	sleeptime = int(raw_input("Enter a sleep time in seconds. Default is zero: ") or "0")
	if type(sleeptime) == int:
		BPTool()
	else:
		print "Sleep time must be a number"
elif proceed == "no" or proceed == "n": 
	quit()				
			
ws.title = 'BP Report for %s' % cust
ws['A1'] = 'Best Practice Report for %s - %s' % (cust, curdate)
#ws.auto_filter.add_filter_column(0, "Pass", blank=False)
ws.conditional_formatting.add('E3:E1048576', FormulaRule(formula=['NOT(ISERROR(SEARCH("Pass",E3)))'], stopIfTrue=True, fill=greenFill))	
ws.conditional_formatting.add('E3:E1048576', FormulaRule(formula=['NOT(ISERROR(SEARCH("Fail",E3)))'], stopIfTrue=True, fill=redFill))				
wb.save('bp-results-' + cust + '-' + str(repdate) + '.xlsx') 
print ""
print "##############################################################"
print "Thank you for using the Palo Alto Best Practices Analysis Tool." 
print "Your output file should be in the directory in which you launched the tool."
print "May the force be with you."
